import os

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .utils import MASTER_XLSX, logger, get_number_of_projects
from ..utils import ROOT_PATH, runtime_config, CONFIG_FILE

runtime_config.read(CONFIG_FILE)


def process_master(source_wb, project_number):
    """
    Function which is called on each cycle in main loop. Takes a master workbook
    and a project number as arguments. Creates a new workbook, populates it with
     the required data from the source_wb file passed in, formats it, then returns
     the workbook from the function, along with the project name which is used
     to name the file.
    """

    wb = Workbook()
    sheet = wb.active
    ws2 = source_wb.active

    al = Alignment(horizontal="left", vertical="top", wrap_text=True,
                   shrink_to_fit=True)

    al2 = Alignment(horizontal="center", vertical="center", wrap_text=True,
                    shrink_to_fit=True)

    double_bottom_border = Border(left=Side(style='thin'),
                                  right=Side(style='none'),
                                  top=Side(style='none'),
                                  bottom=Side(style='double'))

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    bold18Font = Font(size=18, bold=True)

    project_name = ws2.cell(row=1, column=project_number).value
    SRO_name = ws2.cell(row=59, column=project_number).value
    WLC_value = ws2.cell(row=304, column=project_number).value
    project_stage = ws2.cell(row=281, column=project_number).value
    SRO_conf = ws2.cell(row=57, column=project_number).value
    # SRO_conf_last_qtr =
    SoP = ws2.cell(row=201, column=project_number).value
    finance_DCA = ws2.cell(row=280, column=project_number).value
    benefits_DCA = ws2.cell(row=1152, column=project_number).value
    SRO_Comm = ws2.cell(row=58, column=project_number).value
# red_color = 'ffc7ce'
# red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
# sheet.conditional_formatting.add('B5', CellIsRule(operator='containsText', formula=['Amber/Green'], fill=red_fill))
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 15
    sheet.row_dimensions[1].height = 30
    sheet.merge_cells('A1:F1')
    sheet['A1'] = project_name
    sheet['A1'].font = bold18Font
    sheet['A1'].border = double_bottom_border
    sheet['B1'].border = double_bottom_border
    sheet['C1'].border = double_bottom_border
    sheet['D1'].border = double_bottom_border
    sheet['E1'].border = double_bottom_border
    sheet['F1'].border = double_bottom_border
    sheet['A1'].alignment = al2
    sheet.row_dimensions[2].height = 10
    sheet.row_dimensions[3].height = 20
    sheet['A3'] = 'SRO'
    sheet.merge_cells('C3:F3')
    sheet['C3'] = SRO_name
    sheet.row_dimensions[4].height = 10
    sheet.row_dimensions[5].height = 20
    sheet['A5'] = 'WLC:'
    sheet['B5'] = WLC_value
    sheet['C5'] = 'Project Stage:'
    sheet['D5'] = project_stage
    sheet['E5'] = 'Start of Ops:'
    sheet['F5'] = SoP
    sheet.row_dimensions[6].height = 10
    sheet.row_dimensions[7].height = 20
    sheet['A7'] = 'DCA now'
    sheet['B7'] = SRO_conf
    sheet['B7'].border = thin_border
    sheet['C7'] = 'DCA last quarter'
    #sheet['D7'] = SRO_conf_last_qtr
    sheet['D7'].border = thin_border
    sheet['E7'] = 'IPA DCA'
    sheet['F7'].border = thin_border
    sheet.row_dimensions[8].height = 10
    sheet.row_dimensions[9].height = 20
    sheet['A9'] = 'Finance DCA'
    sheet['B9'] = finance_DCA
    sheet['B9'].border = thin_border
    sheet['C9'] = 'Benefits DCA'
    sheet['D9'] = benefits_DCA
    sheet['D9'].border = thin_border
    sheet.row_dimensions[10].height = 10
    sheet.merge_cells('A11:F40')
    sheet['A11'] = SRO_Comm
    sheet['A11'].alignment = al
    sheet['A11'].border = double_bottom_border
    sheet['B11'].border = double_bottom_border
    sheet['C11'].border = double_bottom_border
    sheet['D11'].border = double_bottom_border
    sheet['E11'].border = double_bottom_border
    sheet['F11'].border = double_bottom_border
    sheet['A40'].border = double_bottom_border
    sheet['B40'].border = double_bottom_border
    sheet['C40'].border = double_bottom_border
    sheet['D40'].border = double_bottom_border
    sheet['E40'].border = double_bottom_border
    sheet['F40'].border = double_bottom_border

    return wb, project_name  # outputs a tuple of (wb, project_name) <- parens are optional!


def run(output_path=None, user_provided_master_path=None):

    if user_provided_master_path:
        logger.info(f"Using master file: {user_provided_master_path}")
        q2 = load_workbook(user_provided_master_path)
    else:
        logger.info(f"Using default master file (refer to config.ini)")
        q2 = load_workbook(MASTER_XLSX)

    # get the number of projects, so we know how many times to loop
    project_count = get_number_of_projects(q2)

    for p in range(2, project_count + 2):  # start at 2, representating col B in master; go until number of projects plus 2

        # pass out master and project number into the process_master() function
        # we capture the workbook object and project name in a tuple (these are the objects passed out by the return statement inside process_master() function
        output_wb, project_name = process_master(q2, p)

        # save the file, using the project_name variable in the file name
        try:
            if output_path:
                output_wb.save(os.path.join(output_path[0], f'{project_name}_ANNEX.xlsx'))
                logger.info(f"{project_name}_ANNEX.xlsx to {output_path}")
            else:
                output_path = os.path.join(ROOT_PATH, 'output')
                output_wb.save(os.path.join(output_path, f'{project_name}_ANNEX.xlsx'))
                logger.info(f"{project_name}_ANNEX.xlsx to {output_path}")
                output_path = ""
        except PermissionError:
            logger.critical(f"Cannot save {project_name}_ANNEX.xlsx file - you already have it open. Close and run again.")
            return


if __name__ == "__main__":
    run()
