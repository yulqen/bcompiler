import copy
import os
import unicodedata

from bcompiler.core import Quarter, Master, Row
from ..utils import logger, ROOT_PATH, CONFIG_FILE, runtime_config
from itertools import zip_longest

from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl import Workbook
from openpyxl.drawing.line import LineProperties

runtime_config.read(CONFIG_FILE)


def _calc_quarter_totals(global_data: dict):
    global_totals = {}
    q1_entries = [item[1][1] for item in iter(global_data.items())]
    q2_entries = [item[1][2] for item in iter(global_data.items())]
    q3_entries = [item[1][3] for item in iter(global_data.items())]
    q4_entries = [item[1][4] for item in iter(global_data.items())]
    global_totals['q1', 'rdel'] = round(sum([item['RDEL Total Forecast'] for item in q1_entries]), 2)
    global_totals['q1', 'cdel'] = round(sum([item['CDEL Total Forecast'] for item in q1_entries]), 2)
    global_totals['q1', 'non-gov'] = round(sum([item['Non-Gov Total Forecast'] for item in q1_entries]), 2)
    global_totals['q1', 'total_forecast'] = round(sum([item['Total Forecast'] for item in q1_entries]), 2)
    global_totals['q1', 'total_forecast_sr'] = round(sum([item['Total Forecast SR (20/21)'] for item in q1_entries]), 2)

    global_totals['q2', 'rdel'] = round(sum([item['RDEL Total Forecast'] for item in q2_entries]), 2)
    global_totals['q2', 'cdel'] = round(sum([item['CDEL Total Forecast'] for item in q2_entries]), 2)
    global_totals['q2', 'non-gov'] = round(sum([item['Non-Gov Total Forecast'] for item in q2_entries]), 2)
    global_totals['q2', 'total_forecast'] = round(sum([item['Total Forecast'] for item in q2_entries]), 2)
    global_totals['q2', 'total_forecast_sr'] = round(sum([item['Total Forecast SR (20/21)'] for item in q2_entries]), 2)

    global_totals['q3', 'rdel'] = round(sum([item['RDEL Total Forecast'] for item in q3_entries]), 2)
    global_totals['q3', 'cdel'] = round(sum([item['CDEL Total Forecast'] for item in q3_entries]), 2)
    global_totals['q3', 'non-gov'] = round(sum([item['Non-Gov Total Forecast'] for item in q3_entries]), 2)
    global_totals['q3', 'total_forecast'] = round(sum([item['Total Forecast'] for item in q3_entries]), 2)
    global_totals['q3', 'total_forecast_sr'] = round(sum([item['Total Forecast SR (20/21)'] for item in q3_entries]), 2)

    global_totals['q4', 'rdel'] = round(sum([item['RDEL Total Forecast'] for item in q4_entries]), 2)
    global_totals['q4', 'cdel'] = round(sum([item['CDEL Total Forecast'] for item in q4_entries]), 2)
    global_totals['q4', 'non-gov'] = round(sum([item['Non-Gov Total Forecast'] for item in q4_entries]), 2)
    global_totals['q4', 'total_forecast'] = round(sum([item['Total Forecast'] for item in q4_entries]), 2)
    global_totals['q4', 'total_forecast_sr'] = round(sum([item['Total Forecast SR (20/21)'] for item in q4_entries]), 2)

    return global_totals


def _replace_underscore(name: str):
    return name.replace('/', '_')


def _color_gen():
    for c in [
        'ce5089',
        'ce5650',
        'ce50c8',
        '5050ce',
        '8f50ce',
        '508fce',
        '50ceac',
        '50b1ce',
        '50ce6d'
    ]:
        yield c


def _create_chart(worksheet):
    """Create the fucking chart"""
    chart = ScatterChart()
    chart.varyColors = True
    chart.title = "Financial Analysis"
    chart.style = 1
    chart.height = 10
    chart.width = 20
    chart.x_axis.title = "Financial Quarter"
    chart.y_axis.title = "Cost"
    chart.legend = None
    chart.x_axis.majorUnit = 0.5
    chart.x_axis.minorGridlines = None
#   chart.y_axis.majorUnit = 200

    xvalues = Reference(worksheet, min_col=1, min_row=3, max_row=6)
    picker = _color_gen()
    for i in range(2, 7):
        values = Reference(worksheet, min_col=i, min_row=2, max_row=6)
        series = Series(values, xvalues, title_from_data=True)
        series.smooth = True
        series.marker.symbol = "circle"
        lineProp = LineProperties(solidFill=next(picker))
        series.graphicalProperties.line = lineProp
        chart.series.append(series)
    worksheet.add_chart(chart, "G1")
    return worksheet


def run(masters_repository_dir, output_path=None):

    q1 = Quarter(1, 2017)
    q2 = Quarter(2, 2017)
    q3 = Quarter(3, 2016)
    q4 = Quarter(4, 2016)

    # TODO - we need a function in here that gleans quarter from the filename
    # of the master

    master_q1 = Master(q1, os.path.join(masters_repository_dir, runtime_config['AnalyserFinancialAnalysis']['q1_master']))
    master_q2 = Master(q2, os.path.join(masters_repository_dir, runtime_config['AnalyserFinancialAnalysis']['q2_master']))
    master_q3 = Master(q3, os.path.join(masters_repository_dir, runtime_config['AnalyserFinancialAnalysis']['q3_master']))
    master_q4 = Master(q4, os.path.join(masters_repository_dir, runtime_config['AnalyserFinancialAnalysis']['q4_master']))
#   target_keys = [
#       'RDEL Total Forecast',
#       'CDEL Total Forecast',
#       'Non-Gov Total Forecast',
#       'Total Forecast',
#       'Total Forecast SR (20/21)'
#   ]

    target_keys = runtime_config['AnalyserFinancialAnalysis']['target_keys'].split('\n')

    q3_keys = runtime_config['AnalyserFinancialAnalysis']['q3_target_keys'].split('\n')
    q4_keys = runtime_config['AnalyserFinancialAnalysis']['q4_target_keys'].split('\n')

    # projects from latest master
    projects = master_q2.projects

    project_totals = {key: t for key in target_keys for t in [0]}
    # issue is here - in using a single dict in a list
    project_totals = {q: copy.copy(pt) for q in range(1, 5) for pt in [project_totals]}
    global_totals = {}

    def _update_total(keys: list, target_keys: list, data: list, quarter=None):
        keys, target_keys = target_keys, keys
#       z = list(zip_longest(project_totals['1'].keys(), data)) # don't like the hardcode here in the key
        z = list(zip_longest(keys, data))
        for t in z:
            try:
                project_totals[quarter][t[0]] += t[1]
            except TypeError:
                pass



    # set up sheets
    for p in projects:
        wb = Workbook()
        ws = wb.active
        start_row = 1
        ws.cell(row=start_row, column=1, value=p)
        header = Row(2, start_row + 1, target_keys)
        header.bind(ws)

        for m in [master_q3, master_q4, master_q1, master_q2]:
            try:
                p_data = m[p]
            except KeyError:
                logger.warning(f"Cannot find {p}")
                continue
            if m.quarter.quarter == 1:
                d = p_data.pull_keys(target_keys, flat=True)
                _update_total(target_keys, target_keys, d, m.quarter.quarter)
            if m.quarter.quarter == 2:
                d = p_data.pull_keys(target_keys, flat=True)
                _update_total(target_keys, target_keys, d, m.quarter.quarter)
            if m.quarter.quarter == 3:
                d = p_data.pull_keys(q3_keys, flat=True)
                _update_total(q3_keys, target_keys, d, m.quarter.quarter)
            elif m.quarter.quarter == 4:
                d = p_data.pull_keys(q4_keys, flat=True)
                _update_total(q4_keys, target_keys, d, m.quarter.quarter)
            ws.cell(row=start_row + 2, column=1, value=str(m.quarter))
            r = Row(2, start_row + 2, d)
            r.bind(ws)

            start_row += 1

        global_totals[p] = project_totals
        project_totals = {key: t for key in target_keys for t in [0]}
        project_totals = {q: copy.copy(pt) for q in range(1, 5) for pt in [project_totals]}

        _create_chart(ws)

        if output_path:
            wb.save(os.path.join(output_path[0], f'{p}_FINANCIAL_ANALYSIS.xlsx'))
            logger.info(f"Saved {p}_FINANCIAL_ANALYSIS.xlsx to {output_path}")
        else:
            output_path = os.path.join(ROOT_PATH, 'output')
            wb.save(os.path.join(output_path, f'{p}_FINANCIAL_ANALYSIS.xlsx'))
            logger.info(f"Saved {p}_FINANCIAL_ANALYSIS.xlsx to {output_path}")
            output_path = None


    tots = _calc_quarter_totals(global_totals)

    wb = Workbook()
    ws = wb.active
    start_row = 1
    ws.cell(row=start_row, column=1, value='Totals')
    header = Row(2, start_row + 1, target_keys)
    header.bind(ws)

    for q in ["Q3", "Q4", "Q1", "Q2"]:
        ws.cell(row=start_row + 2, column=1, value=q)
        start_row += 1

    q1s = [t[1] for t in iter(tots.items()) if t[0][0] == 'q1']
    q2s = [t[1] for t in iter(tots.items()) if t[0][0] == 'q2']
    q3s = [t[1] for t in iter(tots.items()) if t[0][0] == 'q3']
    q4s = [t[1] for t in iter(tots.items()) if t[0][0] == 'q4']
    start_row = 1
    for i in [q3s, q4s, q1s, q2s]:
        Row(2, start_row + 2, i).bind(ws)
        start_row += 1

    _create_chart(ws)

    if output_path:
        wb.save(os.path.join(output_path[0], f'TOTAL_FINANCIAL_ANALYSIS.xlsx'))
        logger.info(f"Saved TOTAL_FINANCIAL_ANALYSIS.xlsx to {output_path}")
    else:
        output_path = os.path.join(ROOT_PATH, 'output')
        wb.save(os.path.join(output_path, f'TOTAL_FINANCIAL_ANALYSIS.xlsx'))
        logger.info(f"Saved TOTAL_FINANCIAL_ANALYSIS.xlsx to {output_path}")
        output_path = None




if __name__ == '__main__':
    run(ROOT_PATH)
