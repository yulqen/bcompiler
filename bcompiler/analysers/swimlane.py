import os
import logging

from typing import Tuple

import openpyxl

from ..utils import ROOT_PATH

# typing imports
import sys
from openpyxl.worksheet.worksheet import Worksheet

import datetime
from openpyxl.chart import ScatterChart, Reference, Series

MASTER_XLSX = 'Q2_1718_master.xlsx'

logger = logging.getLogger('bcompiler.compiler')

HOME = os.path.abspath(os.path.expanduser('~'))
DESKTOP = os.path.join(HOME, 'Desktop')

NUMBER_OF_PROJECTS = 32


def gather_data(
        start_row: int,
        project_number: int,
        newwb: openpyxl.Workbook,
        block_start_row: int = 90,
        interested_range: int = 365):
    """
    Gather data from
    :type int: start_row
    :type int project_number
    :type openpyxl.Workbook: newwb
    :type int: block_start_row
    :type int: interested_range
    :rtype: Tuple
    """
    newsheet: Worksheet = newwb.active
    col = project_number + 1
    start_row = start_row + 1

    wb = openpyxl.load_workbook(
        os.path.join(
            DESKTOP, MASTER_XLSX))
    sheet = wb.active

    # print project title
    newsheet.cell(row=start_row - 1, column=1, value=sheet.cell(row=1, column=col).value)
    logger.info(f"Processing: {sheet.cell(row=1, column=col).value}")

    x = start_row
    for i in range(block_start_row, 269, 6):
        val = sheet.cell(row=i, column=col).value
        newsheet.cell(row=x, column=1, value=val)
        x += 1
    x = start_row
    for i in range(block_start_row + 1, 270, 6):
        val = sheet.cell(row=i, column=col).value
        newsheet.cell(row=x, column=2, value=val)
        x += 1

    today = datetime.datetime.today()
    current_row = start_row
    for i in range(91, 269, 6):
        time_line_date = sheet.cell(row=i, column=col).value
        try:
            difference = (time_line_date - today).days
            if difference in range(1, interested_range):
                newsheet.cell(row=current_row, column=3, value=difference)
        except TypeError:
            pass
        finally:
            current_row += 1

    for i in range(start_row, start_row + 30):
        newsheet.cell(row=i, column=4, value=project_number)

    return newwb, start_row


def _segment_series() -> Tuple:
    """Generator for step value when stepping through rows within a project block."""
    cut = dict(
        sobc=1,
        obc=1,
        ds1=4,
        fbc=1,
        ds2=4,
        ds3=4,
        free=8
    )
    for item in cut.items():
        yield item


def _series_producer(sheet, start_row: int, step: int) -> Tuple[Series, int]:
    """
    Generates a single Series() object, containing a Reference() object for x and y values for the chart.
    Implemented as part of a loop; also returns new_start which is the row number it should continue with
    on the next loop.
    :type sheet: Worksheet
    :type start_row: int
    :type step: int
    :return: tuple of items from cut
    """
    xvalues = Reference(sheet, min_col=3, min_row=start_row, max_row=start_row + step)
    values = Reference(sheet, min_col=4, min_row=start_row, max_row=start_row + step)
    series = Series(values, xvalues)
    new_start = start_row + step + 1
    return series, new_start


def _row_calc(project_number: int) -> Tuple[int, int]:
    """
    Helper function to calculate row numbers when writing column of project values to cols A & B.
    :type project_number: int
    :return:  tuple of form (project_number, calculated rows in project block)
    """
    if project_number == 1:
        return 1, 1
    if project_number == 2:
        return 2, 32
    else:
        return (project_number, (project_number + 30) + ((project_number - 2) * 30))


def run(output_path=None):
    wb = openpyxl.Workbook()
    segment_series_generator = _segment_series()
    for p in range(1, 31):
        proj_num, st_row = _row_calc(p)
        wb = gather_data(st_row, proj_num, wb, block_start_row=90, interested_range=365)[0]

    chart = ScatterChart()
    chart.title = "Swimlane Chart"
    chart.style = 1
    chart.x_axis.title = 'Days from Today'
    chart.y_axis.title = 'Project No'
    chart.legend = None
    chart.x_axis.majorUnit = 50
    chart.x_axis.minorGridlines = None
    chart.y_axis.majorUnit = 1

    derived_end = 2

    for p in range(1, NUMBER_OF_PROJECTS):
        for i in range(1,
                       8):  # 8 here is hard-coded number of segments within a project series (ref: dict in _segment_series()
            if i == 1:
                inner_start_row = derived_end
            else:
                inner_start_row = derived_end
            _inner_step = next(segment_series_generator)[1]
            series, derived_end = _series_producer(wb.active, inner_start_row, _inner_step)
            if _inner_step == 1:
                series.marker.symbol = "triangle"
                series.marker.graphicalProperties.solidFill = "01a852"
            else:
                series.marker.symbol = "square"
                series.marker.graphicalProperties.solidFill = "FF0000"
            series.marker.size = 10
            chart.series.append(series)
        segment_series_generator = _segment_series()
        derived_end = derived_end + 1

    wb.active.add_chart(chart, "E1")
    try:
        if output_path:
            wb.save(os.path.join(output_path[0], 'swimlane.xlsx'))
            logger.info(f"Saved swimlane.xlsx to {output_path}")
        else:
            output_path = os.path.join(ROOT_PATH, 'output')
            wb.save(os.path.join(output_path, 'swimlane.xlsx'))
            logger.info(f"Saved swimlane.xlsx to {output_path}")
    except PermissionError:
        logger.critical("Cannot save output.xlsx file - you already have it open. Close and run again.")
        return


if __name__ == "__main__":
    run()
