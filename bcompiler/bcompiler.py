import argparse
import csv
import os
import re
import shutil

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation



def get_parser():
    parser = argparse.ArgumentParser(description='Compile BICC data or prepare Excel BICC return forms.')
    parser.add_argument('-c', '--clean-datamap', dest='datamap', metavar='datamap file', nargs=1,
                        help='clean datamap file'
                             'whose path is given as string')
    parser.add_argument('-v', '--version', help='displays the current version of bcompiler', action="store_true")
    parser.add_argument('-p', '--parse', dest='parse', metavar='source file', nargs=1, help='parse master.csv and flip'
                                                                                            ' to correct orientation')
    parser.add_argument('-b', '--populate-blank', dest='populate', metavar='project integer',
                        help='populate blank bicc forms from master for project N')
    parser.add_argument('-a', '--all', action="store_true")
    parser.add_argument('-d', '--create-wd', dest='create-wd', action="store_true",
                        help='create working directory at $HOME/Documents/bcomiler')
    parser.add_argument('-f', '--force-create-wd', dest='f-create-wd', action="store_true", help='remove existing '
                                                                                                 'working directory and'
                                                                                                 'create a new one')
    return parser

def _clean_datamap(source_file):

    CLEANED_DATAMAP_FILE = 'source_files/cleaned_datamap'
    try:
        os.remove(CLEANED_DATAMAP_FILE)
    except FileNotFoundError:
        pass
    cleaned_datamap = open(CLEANED_DATAMAP_FILE, 'a+')
    with open(source_file, 'r', encoding='UTF-8') as f:
        # make sure every line has a comma at the end
        for line in f.readlines():
            newline = line.rstrip()
            if ',' in newline[-1]:
                newline = newline + '\n'
                cleaned_datamap.write(newline)
            else:
                newline = newline + ',' + '\n'
                cleaned_datamap.write(newline)


def _parse_csv_to_file(source_file):
    """
    Transposes the master to a new master_transposed.csv file.
    :param source_file:
    :return:
    """
    output = open('source_files/master_transposed.csv', 'w+')
    with open(source_file, 'r') as source_f:
        lis = [x.split(',') for x in source_f]
        for i in lis:
            # we need to do this to remove trailing "\n" from the end of each original master.csv line
            i[-1] = i[-1].rstrip()

    for x in zip(*lis):
        for y in x:
            output.write(y + ',')
        output.write('\n')


def create_master_dict_transposed(source_master_csv):
    _parse_csv_to_file(source_master_csv)
    with open('source_files/master_transposed.csv', 'r') as f:
        r = csv.DictReader(f)
        l = [row for row in r]
    return l


def _get_list_projects(source_master_file):
    reader = create_master_dict_transposed(source_master_file)
    pl = [row['Project/Programme Name'] for row in reader]
    return pl

def get_datamap():
    cell_regex = re.compile('[A-Z]+[0-9]+')
    dropdown_headers = _get_dropdown_headers()
    output_excel_map_list = []
    f = open('source_files/cleaned_datamap', 'r')
    data = f.readlines()
    for line in data:
        # split on , allowing us to access useful data from data map file
        data_map_line = line.split(',')
        if data_map_line[1] in ['Summary', 'Finance & Benefits', 'Resources', 'Approval & Project milestones',
                                'Assurance planning']:
            # the end item in the list is a newline - get rid of that
            del data_map_line[-1]
        if cell_regex.search(data_map_line[-1]):
            try:
                m_map = dict(cell_description=data_map_line[0],
                             sheet=data_map_line[1],
                             cell_coordinates=data_map_line[2],
                             validation_header='')
            except IndexError:
                m_map = dict(cell_description=data_map_line[0],
                             sheet="CAN'T FIND SHEET")
            output_excel_map_list.append(m_map)
        elif data_map_line[-1] in dropdown_headers:
            try:
                m_map = dict(cell_description=data_map_line[0],
                             sheet=data_map_line[1],
                             cell_coordinates=data_map_line[2],
                             validation_header=data_map_line[3]
                             )
            except IndexError:
                print("Something wrong with the datamap indexing", m_map.items())

            output_excel_map_list.append(m_map)

    return output_excel_map_list

def project_data_line():
    dict = {}
    with open('source_files/master_transposed.csv', 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = row.pop('Project/Programme Name')
            if key in dict:
                pass
            dict[key] = row
    return dict


def populate_blank_bicc_form(source_master_file, proj_num):
    datamap = get_datamap()
    proj_data = project_data_line()
    ls = _get_list_projects(source_master_file)
    test_proj = ls[int(proj_num)]
    test_proj_data = proj_data[test_proj]
    blank = load_workbook('source_files/bicc_template.xlsx')
    ws_summary = blank.get_sheet_by_name('Summary')
    ws_fb = blank.get_sheet_by_name('Finance & Benefits')
    ws_res = blank.get_sheet_by_name('Resources')
    ws_apm = blank.get_sheet_by_name('Approval & Project milestones')
    ws_ap = blank.get_sheet_by_name('Assurance planning')
    for item in datamap:
        if item['sheet'] == 'Summary':
            if 'Project/Programme Name' in item['cell_description']:
                ws_summary[item['cell_coordinates']].value = test_proj
            try:
                ws_summary[item['cell_coordinates']].value = test_proj_data[item['cell_description']]
            except KeyError:
                print("Cannot find {} in master.csv".format(item['cell_description']))
                pass
            if item['validation_header']:
                dv = create_validation(item['validation_header'])
                ws_summary.add_data_validation(dv)
                dv.add(ws_summary[item['cell_coordinates']])
        elif item['sheet'] == 'Finance & Benefits':
            try:
                ws_fb[item['cell_coordinates']].value = test_proj_data[item['cell_description']]
            except KeyError:
                print("Cannot find {} in master.csv".format(item['cell_description']))
                pass
        elif item['sheet'] == 'Resources':
            try:
                ws_res[item['cell_coordinates']].value = test_proj_data[item['cell_description']]
            except KeyError:
                print("Cannot find {} in master.csv".format(item['cell_description']))
                pass
        elif item['sheet'] == 'Approval & Project milestones':
            try:
                ws_apm[item['cell_coordinates']].value = test_proj_data[item['cell_description']]
            except KeyError:
                print("Cannot find {} in master.csv".format(item['cell_description']))
                pass
        elif item['sheet'] == 'Assurance planning':
            try:
                ws_ap[item['cell_coordinates']].value = test_proj_data[item['cell_description']]
            except KeyError:
                print("Cannot find {} in master.csv".format(item['cell_description']))
                pass

    blank.save('source_files/{}_Q2_Return.xlsx'.format(test_proj))

def pop_all():
    number_of_projects = len(_get_list_projects('source_files/master.csv'))
    for p in range(number_of_projects):
        populate_blank_bicc_form('source_files/master.csv', p)


def _create_working_directory():
    """
    We need a proper directory to work in.
    :return:
    """
    docs = os.path.join(os.path.expanduser('~'), 'Documents')
    bcomp_working_d = 'bcompiler'
    root_path = os.path.join(docs, bcomp_working_d)
    folders = ['source', 'output']
    if not os.path.exists(root_path):
        os.mkdir(root_path)
        for folder in folders:
            os.mkdir(os.path.join(root_path, folder))
        print("Clean working directory created at {}".format(root_path))
    else:
        print("Working directory exists. You can either run the program like this and files"
              "will be overwritten, or you should do --force-create-wd to remove the working"
              "directory and create a new one.\n\nWARNING: this will remove any datamap and master.csv"
              "files persent.")


def _delete_working_directory():
    docs = os.path.join(os.path.expanduser('~'), 'Documents')
    bcomp_working_d = 'bcompiler'
    root_path = os.path.join(docs, bcomp_working_d)
    try:
        shutil.rmtree(root_path)
        return "{} deleted".format(root_path)
    except FileNotFoundError:
        return


def _get_dropdown_data(header=None):
    """
    Pull the dropdown data from the Dropdown List sheet in bicc_template.xlsx. Location
    of this template file might need to be dynamic.
    :return tuple of column values from sheet, with header value at list[0]:
    """
    wb = load_workbook('source_files/bicc_template.xlsx', data_only=True)
    ws = wb.get_sheet_by_name('Dropdown List')
    columns = ws.columns
    col_lis = [col for col in columns]
    dropdown_data = [[c.value for c in t if c.value] for t in col_lis]
    if header:
        h = [h for h in dropdown_data if header in h[0]]
        h = tuple(h[0])
        print("Getting {}".format(h))
        return h
    else:
        return dropdown_data

def _get_dropdown_headers():
    wb = load_workbook('source_files/bicc_template.xlsx', data_only=True)
    ws = wb.get_sheet_by_name('Dropdown List')
    rows = ws.rows
    a_row = next(rows)
    return [h.value for h in a_row]

def create_validation(header):
    t = _get_dropdown_data(header)
    t = t[1:]
    t_str = ",".join(map(str, t))
    dv = DataValidation(type='list', formula1=t_str, allow_blank=True)
    dv.prompt = "Please select from the list"
    dv.promptTitle = 'List Selection'
    return dv


# Validation data TODO this is perfect for a thread
#VAL_QUARTER = _get_dropdown_data(header='Quarter')
#VAL_JOINING_QTR = _get_dropdown_data(header='Joining Qtr')
#VAL_CLASSIFICATION = _get_dropdown_data(header='Classification')
#VAL_AGENCIES = _get_dropdown_data(header='Agencies')
#VAL_GROUP = _get_dropdown_data(header='Group')
#VAL_DFT_DIVISION = _get_dropdown_data(header='DfT Division')
#VAL_ENTITY = _get_dropdown_data(header='Entity')
#VAL_METHODOLOGY = _get_dropdown_data(header='Methodology')
#VAL_CATEGORY = _get_dropdown_data(header='Category')
#VAL_SCOPE_CHANGED = _get_dropdown_data(header='Scope Changed')
#VAL_MONETISED = _get_dropdown_data(header='Monetised / Non Monetised Benefits')
#VAL_SDP = _get_dropdown_data(header='SDP')
# TODO continue with this

def main():
    parser = get_parser()
    args = vars(parser.parse_args())
    if args['version']:
        print("1.0")
        return
    if args['datamap']:
        _clean_datamap(args['datamap'][0])
        print("{} cleaned".format(args['datamap'][0]))
        return
    if args['parse']:
        _parse_csv_to_file(args['parse'][0])
        return
    if args['populate']:
        _clean_datamap('source_files/datamap')
        _parse_csv_to_file('source_files/master.csv')
        populate_blank_bicc_form('source_files/master.csv', args['populate'])
        return
    if args['all']:
        pop_all()
        return
    if args['create-wd']:
        _create_working_directory()
        return
    if args['f-create-wd']:
        print("This will destroy your existing working directory prior to creating a new one.\n\nAre you sure?")
        response = input('(y/n) --> ')
        if response in ('y', 'ye', 'yes', 'Y', 'YES'):
            _delete_working_directory()
            _create_working_directory()
            return
        else:
            return

if __name__ == '__main__':
    _get_dropdown_data('Quarter')
    main()