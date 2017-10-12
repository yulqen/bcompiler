import decimal
import fnmatch
import logging
import os
import re
import sys

from datetime import date, datetime
from typing import Dict, List

from openpyxl import load_workbook, Workbook

from bcompiler.process import Cleanser
from bcompiler.process.cellformat import CellFormatState
from bcompiler.process.simple_comparitor import FileComparitor, ParsedMaster
from bcompiler.utils import DATAMAP_RETURN_TO_MASTER, OUTPUT_DIR, RETURNS_DIR
from bcompiler.utils import runtime_config as config
from bcompiler.process.cleansers import DATE_REGEX_TIME
from bcompiler.process.datamap import Datamap

CELL_REGEX = re.compile('[A-Z]+[0-9]+')
DROPDOWN_REGEX = re.compile('^\D*$')
TODAY = date.today().isoformat()

logger = logging.getLogger('bcompiler.compiler')

DATA_MAP_FILE = DATAMAP_RETURN_TO_MASTER


def get_current_quarter(source_file: str) -> str:
    """
    :param source_file: path to a template file whose G3 cell contains the \
    current quarter.
    :return: str of the current quarter
    :rtype: str

    .. note::

        This is a note.

    .. warning::
        This cell should not be hard-coded. Move to config.ini.

    """
    wb = load_workbook(RETURNS_DIR + source_file, read_only=True)
    ws = wb[config['TemplateSheets']['summary_sheet']]
    q = ws['G3'].value
    logger.info('Getting current Quarter as {}'.format(q))
    return q


def encode_win(s: str):
    """
    Will take cp1252 encoded, covert to bytes and re-encode in utf-8
    """
    return s.encode('cp1252').decode('utf-8')


def parse_source_cells(source_file: str, datamap_source_file: str) -> \
        List[Dict[str, str]]:
    """
    Takes an Excel source file (populated template), and a datamap and extracts
    the data from the source file according to datamap mappings.
    """
    ls_of_dataline_dicts = []
    wb = load_workbook(source_file, read_only=True, data_only=True)
    datamap_obj = Datamap()
    datamap_obj.cell_map_from_csv(datamap_source_file)
    for item in datamap_obj.cell_map:
        if item.template_sheet is not None and item.cell_reference is not None:
            if os.name == 'nt':
                item.cell_key = encode_win(item.cell_key)
            try:
                ws = wb[item.template_sheet]
            except KeyError as e:
                logger.critical(f"{e}.{source_file} is not a BICC template. Not processing. Remove it!")
                sys.exit()
            try:
                v = ws[item.cell_reference].value
            except IndexError:
                logger.error(
                    "Datamap wants sheet: {}; cellref: {} but this is out"
                    "of range.\n\tFile: {}".format(
                        item.template_sheet,
                        item.cell_reference,
                        source_file))
                v = ""
            except ValueError as e:
                logger.critical(f"{e.args[0]} at cell {ws.title}:{item.cell_reference}. This value will NOT be transferred to master. Skipping...")
                v = "NOT TRANSFERRED DUE TO ERROR: Refer to bcompiler log"
                pass
            else:
                if v is None:
                    logger.debug(
                        "{} in {} is empty.".format(
                            item.cell_reference,
                            item.template_sheet))
                elif type(v) == str:
                    v = v.rstrip()
                elif type(v) == float:
                    v = decimal.Decimal(v)
                    v = v.quantize(decimal.Decimal('.01'), rounding=decimal.ROUND_HALF_EVEN)
                else:
                    logger.debug(
                        "{} in {} is {}".format(
                            item.cell_reference,
                            item.template_sheet,
                            v))
                try:
                    c = Cleanser(v)
                except IndexError:
                    logger.error(
                        ("Trying to clean an empty cell {} at sheet {} in {}. "
                         "Ignoring.").format(
                            item.cell_reference, item.template_sheet, source_file))
                except TypeError:
                    pass
                else:
                    v = c.clean()
            destination_kv = dict(gmpp_key=(item.cell_key).rstrip(), gmpp_key_value=v)
            ls_of_dataline_dicts.append(destination_kv)
    return ls_of_dataline_dicts


def write_excel(source_file, count, workbook, compare_master=None) -> None:
    """
    Writes all return data to a single master Excel sheet.
    """
    ws = workbook.active

    # give it a title
    ws.title = "Constructed BICC Data Master"

    # this is the data from the source spreadsheet
    out_map = parse_source_cells(source_file, DATAMAP_RETURN_TO_MASTER)

    # we need to the project name to work out index order for comparing
    # master file
    project_name = [
        item['gmpp_key_value']
        for item in out_map if item['gmpp_key'] == 'Project/Programme Name'][0]

    try:
        if compare_master:
            compare_file = compare_master[0]
    except TypeError:
        compare_file = None

    if compare_master:
        parsed_master = ParsedMaster(compare_file)
        hd_indices = parsed_master._project_header_index
        try:
            this_index = [
                v for k, v in hd_indices.items() if k == project_name][0]
        except IndexError:
            logger.warning(
                ("Cannot find project title '{}' in previous master. Will "
                 "compile data but not compare until return form and master "
                 "match. Alternatively, this could be a new file.").format(
                    project_name))

    try:
        # this deals with issue of not passing --compare to compile argument
        comparitor = FileComparitor([compare_file])
    except UnboundLocalError:
        pass

    if count == 1:
        i = 1
        # this one writes the first column, the keys
        for d in out_map:
            c = ws.cell(row=i, column=1)
            c.value = d['gmpp_key']
            i += 1
        i = 1

        # then it writes the second column with the values
        for d in out_map:

            c = ws.cell(row=i, column=2)
            c_format = CellFormatState()

            try:
                compare_val = comparitor.compare(this_index, d['gmpp_key'].rstrip())
                if isinstance(compare_val, str) and compare_val is not None and re.match(DATE_REGEX_TIME, compare_val):
                    ds = compare_val.split(' ')
                    comps = [int(x) for x in ds[0].split('-')]
                    compare_val = datetime(*comps)

            except UnboundLocalError:
                compare_val = False

            # TODO - apply number format WITHOUT a compare_val

            # if there is something to compare it
            if compare_val:
                c_format.action(
                    compare_val=compare_val,
                    this_val=d['gmpp_key_value'],
                    key=d['gmpp_key'])
                formt = c_format.export_rule()
                c.fill = formt[0]
                if formt[1] == "":
                    c.value = d['gmpp_key_value']
                else:
                    c.value = d['gmpp_key_value']
                    c.number_format = formt[1]
            else:
                # there is nothing to compare to so no formatting required
                # just print the value
                c.value = d['gmpp_key_value']
            i += 1
    else:
        i = 1
        # now we have no need of the keys any more so we're just writing
        # values here
        for d in out_map:
            c = ws.cell(row=i, column=count + 1)
            c_format = CellFormatState()

            try:
                compare_val = comparitor.compare(this_index, d['gmpp_key'])
                if isinstance(compare_val, str) and compare_val is not None and re.match(DATE_REGEX_TIME, compare_val):
                    ds = compare_val.split(' ')
                    comps = [int(x) for x in ds[0].split('-')]
                    compare_val = datetime(*comps)
            except UnboundLocalError:
                compare_val = False

            # if there is something to compare it
            if compare_val:
                c_format.action(
                    compare_val=compare_val,
                    this_val=d['gmpp_key_value'],
                    key=d['gmpp_key'])
                formt = c_format.export_rule()
                c.fill = formt[0]
                if formt[1] == "":
                    c.value = d['gmpp_key_value']
                else:
                    c.value = d['gmpp_key_value']
                    c.number_format = formt[1]
            else:
                # there is nothing to compare to so no formatting required
                # just print the value
                c.value = d['gmpp_key_value']
            i += 1


def run(compare_master=None):
    """
    Run the compile function.
    """
    # if we want to do a comparison
    if compare_master:

        workbook = Workbook()
        count = 1
        for file in os.listdir(RETURNS_DIR):
            if fnmatch.fnmatch(file, '*.xlsm') or fnmatch.fnmatch(file, '*.XLSX') or fnmatch.fnmatch(file, '*.xlsx'):
                logger.info("Processing {}".format(file))
                write_excel(
                    (os.path.join(RETURNS_DIR, file)),
                    count=count,
                    workbook=workbook,
                    compare_master=compare_master
                )
                count += 1
            elif fnmatch.fnmatch(file, '*.xlsm#' or fnmatch.fnmatch(file, '*.xlsx#')):
                logger.warning("You have a file open in your spreadsheet program. Ignoring the lock file.")
            else:
                logger.critical("There are no Excel files in {}. Copy some in there!".format(RETURNS_DIR))
        OUTPUT_FILE = '/'.join([OUTPUT_DIR, 'compiled_master_{}_{}.xlsx'.format(TODAY, "Q2")])
        workbook.save(OUTPUT_FILE)
    else:
        # we just want a straight master with no change indication
        workbook = Workbook()
        count = 1
        try:
            d = os.listdir(RETURNS_DIR)
        except FileNotFoundError:
            logger.critical("There is no 'returns' directory and therefore "
                            "therefore no"
                            " returns to compile. Ensure you have the path "
                            "'bcompiler/source/returns' and dump your returns "
                            " files in there. bcompiler -d may help.")
            return
        for file in d:
            if fnmatch.fnmatch(file, '*.xlsm') or fnmatch.fnmatch(file, '*.XLSX') or fnmatch.fnmatch(file, '*.xlsx'):
                logger.info("Processing {}".format(file))
                write_excel(
                    (os.path.join(RETURNS_DIR, file)),
                    count=count,
                    workbook=workbook,
                )
                count += 1
            elif fnmatch.fnmatch(file, '*.xlsm#' or fnmatch.fnmatch(file, '*.xlsx#')):
                logger.warning("You have a file open in your spreadsheet program. Ignoring the lock file.")
            else:
                logger.critical("There are no Excel files in {}. Copy some in there!".format(RETURNS_DIR))
        q_string = config['QuarterData']['CurrentQuarter'].split()[0]
        OUTPUT_FILE = '/'.join([OUTPUT_DIR, 'compiled_master_{}_{}.xlsx'.format(TODAY, q_string)])
        workbook.save(OUTPUT_FILE)
