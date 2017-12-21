import re
import datetime
import logging
import unicodedata
from pathlib import Path
from typing import List, Tuple, Iterable, Optional, Any

from ..utils import project_data_from_master
from ..process.cleansers import DATE_REGEX_4

from openpyxl import load_workbook

logger = logging.getLogger('bcompiler.utils')


class MasterProjectData:
    """
    Light wrapper around Master.data object.
    """
    def __init__(self, d: dict) -> None:
        """
        ordered_dict is easiest to get from project_data_from_master[x]
        """
        self._data = d

    def __len__(self) -> int:
        return len(self._data)

    def __getitem__(self, item):
        return self._data[item]

    def key_filter(self, key: str) -> List[Tuple]:
        """
        Return a list of (k, v) tuples if k in master key.
        """
        data = [item for item in self._data.items() if key in item[0]]
        if not data:
            raise KeyError("Sorry, there is no matching data")
        return (data)

    def pull_keys(self, input_iter: Iterable, flat=False) -> List[Tuple[Any, ...]]:
        """
        Returns a list of (key, value) tuples from MasterProjectData if key matches a
        key. The order of tuples is based on the order of keys passed in the iterable.
        """
        if flat is True:
            # search and replace troublesome EN DASH character
            xs = [item for item in self._data.items()
                  for i in input_iter if item[0].strip().replace(unicodedata.lookup('EN DASH'), unicodedata.lookup('HYPHEN-MINUS')) == i]
            xs = [_convert_str_date_to_object(x) for x in xs]
            ts = sorted(xs, key=lambda x: input_iter.index(x[0].strip().replace(unicodedata.lookup('EN DASH'), unicodedata.lookup('HYPHEN-MINUS'))))
            ts = [item[1] for item in ts]
            return ts
        else:
            xs = [item for item in self._data.items()
                  for i in input_iter if item[0].replace(unicodedata.lookup('EN DASH'), unicodedata.lookup('HYPHEN-MINUS')) == i]
            xs = [item for item in self._data.items()
                  for i in input_iter if item[0] == i]
            xs = [_convert_str_date_to_object(x) for x in xs]
            ts = sorted(xs, key=lambda x: input_iter.index(x[0].replace(unicodedata.lookup('EN DASH'), unicodedata.lookup('HYPHEN-MINUS'))))
            return ts

    def __repr__(self):
        return f"MasterProjectData() - with data: {id(self._data)}"


def _convert_str_date_to_object(d_str: tuple) -> Tuple[str, Optional[datetime.date]]:
    try:
        if re.match(DATE_REGEX_4, d_str[1]):
            try:
                ds = d_str[1].split('-')
                return (d_str[0], datetime.date(int(ds[0]), int(ds[1]), int(ds[2])))
            except TypeError:
                return d_str
        else:
            return d_str
    except TypeError:
        return d_str


class Master:
    """
    Master class.
    """
    def __init__(self, quarter, path: str) -> None:
        self._quarter = quarter
        self.path = path
        self._data = project_data_from_master(self.path)
        self._project_titles = [item for item in self.data.keys()]
        self.year = self._quarter.year

    def __getitem__(self, project_name):
        return MasterProjectData(self._data[project_name])

    @property
    def data(self):
        return self._data

    @property
    def quarter(self):
        return self._quarter

    @property
    def filename(self):
        p = Path(self.path)
        return p.name

    @property
    def projects(self):
        return self._project_titles

    def duplicate_keys(self, to_log=None):
        wb = load_workbook(self.path)
        ws = wb.active
        col_a = next(ws.iter_cols())
        col_a = [item.value for item in col_a]
        seen: set = set()
        uniq = []
        dups: set = set()
        for x in col_a:
            if x not in seen:
                uniq.append(x)
                seen.add(x)
            else:
                dups.add(x)
        if to_log and len(dups) > 0:
            for x in dups:
                logger.warning(f"{self.path} contains duplicate key: \"{x}\". Masters cannot contain duplicate keys. Rename them.")
            return True
        elif to_log and len(dups) == 0:
            logger.info(f"No duplicate keys in {self.path}")
            return False
        elif len(dups) > 0:
            return dups
        else:
            return False

    def __repr__(self):
        return f"Master({self.path}, {self.quarter.quarter}, {self.quarter.year})"
