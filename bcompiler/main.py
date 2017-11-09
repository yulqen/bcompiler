"""
Copyright (c) 2016 Matthew Lemon

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy,  modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the  Software is
furnished to do so, subject to the following conditions:
The above copyright notice and this permission notice shall be included in
all copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS
IN THE SOFTWARE. """

import argparse
import datetime
import logging
import os
import re
import sys
import textwrap
from typing import Dict, List

import colorlog
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet.datavalidation import DataValidation

import bcompiler.compile as compile_returns
from bcompiler import __version__
from bcompiler.process import Cleanser
from bcompiler.process.datamap import Datamap
from bcompiler.utils import (CLEANED_DATAMAP, DATAMAP_MASTER_TO_RETURN,
                             DATAMAP_RETURN_TO_MASTER,
                             OUTPUT_DIR, SOURCE_DIR, VALIDATION_REFERENCES,
                             parse_csv_to_file,
                             working_directory, SHEETS, CURRENT_QUARTER,
                             row_data_formatter,
                             BLANK_TEMPLATE_FN, project_data_from_master)
from bcompiler.utils import runtime_config as config
from bcompiler.utils import directory_has_returns_check

from bcompiler.analysers import swimlane_run
from bcompiler.analysers import swimlane_assurance_run
from bcompiler.analysers import annex_run
from bcompiler.analysers import keyword_run

logger = colorlog.getLogger('bcompiler')
logger.setLevel(logging.DEBUG)


def analyser_args(args, func):
    """
    Helper function to parse commandline arguments related to --analyser flag. Func
    is a runner function defined elsewhere that does the work.
    """
    if args['start_date'] and args['end_date']:
        if args['output'] and not args['master']:  # user stipulates an output directory
            func(args['output'], date_range=[args['start_date'][0], args['end_date'][0]])
            return
        if args['output'] and args['master']:  # user stipulates an output and a target master
            func(args['output'], args['master'][0], date_range=[args['start_date'][0], args['end_date'][0]])
            return
        if args['master'] and not args['output']:  # user stipulates a master but NOT an output directory
            func(user_provided_master_path=args['master'][0], date_range=[args['start_date'][0], args['end_date'][0]])
            return
        else:  # no options supplied - default options applied (saved to bcompiler/output, master from config.ini
            func(date_range=[args['start_date'][0], args['end_date'][0]])
            return
    if args['compare']:
        func(args['compare'][0])
        return
    else:
        if args['output'] and not args['master']:  # user stipulates an output directory
            func(args['output'])
            return
        if args['output'] and args['master']:  # user stipulates an output and a target master
            func(args['output'], args['master'][0])
            return
        if args['master'] and not args['output']:  # user stipulates a master but NOT an output directory
            func(user_provided_master_path=args['master'][0])
            return
        else:  # no options supplied - default options applied (saved to bcompiler/output, master from config.ini
            func()


def keyword_args(args, func):
    """
    Helper function to parse commandline arguments related to --analyser keywords option.
    Func is a runner funtion defined elsewhere that does the work.
    """
    if args['xlsx']:
        if not args['master']:  # user stipulates an output directory
            try:
                func(search_term=args['analyser'][1], xlsx=args['xlsx'])
            except IndexError:
                logger.critical("You need to provide a search term, e.g. '--analyser keyword RAG'")
            return
        if args['master']:  # user stipulates an output and a target master
            try:
                func(user_provided_master_path=args['master'][0], search_term=args['analyser'][1], xlsx=args['xlsx'])
            except IndexError:
                logger.critical("You need to provide a search term, e.g. '--analyser keyword RAG'")
            return
        else:  # no options supplied - default options applied (saved to bcompiler/output, master from config.ini
            try:
                func(search_term=args['analyser'][1], xlsx=args['xlsx'])
            except IndexError:
                logger.critical("You need to provide a search term, e.g. '--analyser keyword RAG'")
            return
    else:
        try:
            func(search_term=args['analyser'][1])
        except IndexError:
            logger.critical("You need to provide a search term, e.g. '--analyser keyword RAG'")
        return



def get_parser():
    parser = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=textwrap.dedent("""
            | |__   ___ ___  _ __ ___  _ __ (_) | ___ _ __
            | '_ \ / __/ _ \| '_ ` _ \| '_ \| | |/ _ \ '__|
            | |_) | (_| (_) | | | | | | |_) | | |  __/ |
            |_.__/ \___\___/|_| |_| |_| .__/|_|_|\___|_|
                                    |_|
            Compile BICC data or prepare Excel BICC return forms."""))


    parser.add_argument(
        '-c',
        '--clean-datamap',
        action="store_true",
        dest="clean-datamap",
        help=("Remove trailing spaces from datamap and ensure ready for running. "
              "Should be no requirement to run manually."))
    parser.add_argument(
        '-v',
        '--version',
        action="store_true",
        help='Displays the current version of bcompiler')
    parser.add_argument(
        '-r',
        '--count-rows',
        dest='count-rows',
        action="store_true",
        help='Count rows in each sheet in each return file in output folder')
    parser.add_argument(
        '--csv',
        action="store_true",
        help='If used with -r, will output to csv file in output directory')
    parser.add_argument(
        '--quiet',
        action="store_true",
        help='If used with -r, will only report differences in row count if they occur.')
    parser.add_argument(
        '-t',
        '--transpose',
        dest='transpose',
        metavar='SOURCE_FILE',
        nargs=1,
        help='Tranpose master.csv and flip to opposite orientation')
    parser.add_argument(
        '-b',
        '--populate-bicc-form',
        dest='populate',
        metavar='PROJECT_INTEGER',
        help='Populate blank bicc forms from master for project N')
    parser.add_argument(
        '-a',
        '--all',
        action="store_true",
        help='Populate blank templates with data from master')
    parser.add_argument(
        'compile',
        help='Compile BICC returns to master (note: this can be omitted)',
        default='compile',
        nargs='?')
    parser.add_argument(
        '--compare',
        nargs=1,
        metavar="PATH_TO_FILE TO COMPARE",
        help=('To be used with compile action; file path to master file '
              'to compare to compiled data'))
    parser.add_argument(
        '-ll',
        '--loglevel',
        type=str,
        choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
        help=('Set the logging level for the console.'
              'The log file is set to DEBUG.'))
    parser.add_argument(
        '--analyser',
        nargs='+',
        help=('Refer to documentation for options for ANALYSER.')
    )
    parser.add_argument(
        '--output',
        nargs=1,
        help=('Path to save the resulting file. Ignored if used without --analyser.'),
        metavar="PATH_TO_DIRECTORY"
    )
    parser.add_argument(
        '--master',
        nargs=1,
        help=('Path to master to be used for analysis. Ignored if used without --analyser.'),
        metavar="PATH_TO_DIRECTORY"
    )
    parser.add_argument(
        '--xlsx',
        nargs=1,
        help=('Path to xlsx file to be used as output for keyword analyser')
    )
    parser.add_argument(
        '--start_date',
        nargs=1,
        help='Start date for milestone range - dd/mm/yy')
    parser.add_argument(
        '--end_date',
        nargs=1,
        help='End date for milestone range - dd/mm/yy')

    return parser


def clean_datamap(dm_file):
    """
    Used for its side-effects only which isn't ideal, but this isn't
    Haskell, so why not?
    """
    logger.info("Cleaning {}.".format(dm_file))
    cleaned_datamap_file = CLEANED_DATAMAP
    try:
        os.remove(cleaned_datamap_file)
    except FileNotFoundError:
        pass
    cleaned_datamap = open(cleaned_datamap_file, 'a+')
    try:
        with open(dm_file, 'r', encoding='UTF-8') as f:  # make sure every line has a comma at the end
            for line in f.readlines():
                newline = line.rstrip()
                if ',' in newline[-1]:
                    newline += '\n'
                    cleaned_datamap.write(newline)
                else:
                    newline = newline + ',' + '\n'
                    cleaned_datamap.write(newline)
        cleaned_datamap.close()
    except UnicodeDecodeError:
        with open(dm_file, 'r', encoding='latin1') as f:  # make sure every line has a comma at the end
            for line in f.readlines():
                newline = line.rstrip()
                if ',' in newline[-1]:
                    newline += '\n'
                    cleaned_datamap.write(newline)
                else:
                    newline = newline + ',' + '\n'
                    cleaned_datamap.write(newline)
        cleaned_datamap.close()


def get_list_projects(source_master_file):
    """
    Returns a list of Project/Programme Names.
    """
    try:
        wb = load_workbook(source_master_file)
    except FileNotFoundError:
        logger.critical("Have you copied the compiled master xlsx file into"
                        " the source directory and named it correctly in config.ini?")
        sys.exit(1)
        return
    ws = wb.active
    return [item.value for item in ws[1][1:] if item.value is not None]


def get_datamap():
    """
    The old-style datamap design, using parsing rather than creating a Datamap
    class.
    """
    cell_regex = re.compile('[A-Z]+[0-9]+')
    dropdown_headers = get_dropdown_headers()
    output_excel_map_list = []
    f = open(SOURCE_DIR + 'cleaned_datamap.csv', 'r')
    data = f.readlines()
    for line in data:
        # split on , allowing us to access useful data from data map file
        data_map_line = line.split(',')
        if data_map_line[1] in SHEETS:
            # the end item in the list is a newline - get rid of that
            del data_map_line[-1]
        if cell_regex.search(data_map_line[-1]):
            try:
                m_map = dict(
                    cell_description=data_map_line[0],
                    sheet=data_map_line[1],
                    cell_coordinates=data_map_line[2],
                    validation_header='')
            except IndexError:
                m_map = dict(
                    cell_description=data_map_line[0],
                    sheet="CAN'T FIND SHEET")
            output_excel_map_list.append(m_map)
        elif data_map_line[-1] in dropdown_headers:
            try:
                m_map = dict(
                    cell_description=data_map_line[0],
                    sheet=data_map_line[1],
                    cell_coordinates=data_map_line[2],
                    validation_header=data_map_line[3])
            except IndexError:
                logger.error("Something wrong with the datamap indexing",
                             m_map.items())
            output_excel_map_list.append(m_map)
    return output_excel_map_list


def has_whiff_of_total(desc: str) -> bool:
    total_conditions = [
        'Total',
        'RDEL Total',
        'CDEL Total',
    ]
    for cond in total_conditions:
        if cond in desc:
            return True
        else:
            return False


def imprint_current_quarter(sheet) -> None:
    """
    Overwrites summary g3 cell.
    """
    sheet['G3'].value = CURRENT_QUARTER


def lock_cells(sheets: list, target_cells: List[Dict]) -> None:
    """
    This function will set projection for each sheet in sheets to True
    and then mark all cells to be protected throughout the spreadsheet.

    Example:

        t_cells = [
            {'ws_summary': 'A20'},
            {'ws_summary': 'B10'}
            ]

        lock_cells([ws_summary], t_cells)
    """

    try:
        for s in sheets:
            s.protection.sheet = True
    except:
        print("Cannot access that sheet")

    prot = Protection(locked=True, hidden=False)
    for d in target_cells:
        for k, v in d.items():
            k[v].protection = prot


def populate_blank_bicc_form(source_master_file, proj_num):
    datamap = Datamap()
    datamap.cell_map_from_csv(os.path.join(SOURCE_DIR, config['Datamap']['name']))
    proj_data = project_data_from_master(source_master_file)
    ls = get_list_projects(source_master_file)
    test_proj = ls[int(proj_num)]
    logger.info("Processing project {}.".format(test_proj))
    test_proj_data = proj_data[test_proj]
    blank = load_workbook(SOURCE_DIR + BLANK_TEMPLATE_FN, keep_vba=True)
    ws_summary = blank[config['TemplateSheets']['summary_sheet']]
    ws_fb = blank[config['TemplateSheets']['fb_sheet']]
    ws_res = blank[config['TemplateSheets']['resource_sheet']]
    ws_apm = blank[config['TemplateSheets']['apm']]
    ws_ap = blank[config['TemplateSheets']['ap']]
    ws_gmpp = blank[config['TemplateSheets']['gmpp']]

    # TODO - flagged for removal

    TARGET_LOCK_CELLS = [
        {ws_summary: 'B5'},
        {ws_summary: 'B6'},
        {ws_summary: 'C6'},
        {ws_summary: 'G3'},
        {ws_summary: 'G5'},
        {ws_summary: 'I3'},
        {ws_apm: 'A9'},
        {ws_apm: 'A10'},
        {ws_apm: 'A11'},
        {ws_apm: 'A12'},
        {ws_apm: 'A13'},
        {ws_apm: 'A14'},
        {ws_apm: 'A15'},
        {ws_apm: 'A16'},
        {ws_apm: 'A17'},
        {ws_apm: 'A18'},
        {ws_apm: 'A19'},
        {ws_ap: 'A8'},
        {ws_ap: 'A9'},
        {ws_ap: 'A10'},
        {ws_ap: 'A11'},
        {ws_ap: 'A12'},
        {ws_ap: 'A13'},
        {ws_ap: 'A14'},
        {ws_ap: 'A15'}
    ]

    for item in datamap.cell_map:
        if item.template_sheet == config['TemplateSheets']['summary_sheet']:
            if 'Project/Programme Name' in item.cell_key:
                ws_summary[item.cell_reference].value = test_proj
                continue
            if isinstance(test_proj_data[item.cell_key], datetime.date):
                ws_summary[item.cell_reference].value = test_proj_data[item.cell_key]
                ws_summary[item.cell_reference].number_format = 'dd/mm/yyyy'
                continue
            try:
                if re.match(r'(\d+/\d+/\d+)', test_proj_data[item.cell_key]):
                    ws_summary[item.cell_reference].value = test_proj_data[item.cell_key]
                    ws_summary[item.cell_reference].number_format = 'dd/mm/yyyy'
            except TypeError:
                pass
            if test_proj_data[item.cell_key] is None:
                continue
            c = Cleanser(str(test_proj_data[item.cell_key]))
            cleaned = c.clean()
            ws_summary[item.cell_reference].value = cleaned
        elif item.template_sheet == config['TemplateSheets']['fb_sheet']:
            if has_whiff_of_total(item.cell_key):
                continue
            if isinstance(test_proj_data[item.cell_key], datetime.date):
                c = Cleanser(str(test_proj_data[item.cell_key]))
                cleaned = c.clean()
                ws_fb[item.cell_reference].value = test_proj_data[item.cell_key]
                ws_fb[item.cell_reference].number_format = 'dd/mm/yyyy'
                continue
            try:
                if re.match(r'(\d+/\d+/\d+)', test_proj_data[item.cell_key]):
                    ws_fb[item.cell_reference].value = test_proj_data[item.cell_key]
                    ws_fb[item.cell_reference].number_format = 'dd/mm/yyyy'
            except TypeError:
                pass
            if test_proj_data[item.cell_key] is None:
                continue
            c = Cleanser(str(test_proj_data[item.cell_key]))
            cleaned = c.clean()
            ws_fb[item.cell_reference].value = cleaned
        elif item.template_sheet == config['TemplateSheets']['resource_sheet']:
            if has_whiff_of_total(item.cell_key):
                continue
            if isinstance(test_proj_data[item.cell_key], datetime.date):
                ws_res[item.cell_reference].value = test_proj_data[item.cell_key]
                ws_res[item.cell_reference].number_format = 'dd/mm/yyyy'
            if test_proj_data[item.cell_key] is None:
                continue
            try:
                if re.match(r'(\d+/\d+/\d+)', test_proj_data[item.cell_key]):
                    ws_res[item.cell_reference].value = test_proj_data[item.cell_key]
                    ws_res[item.cell_reference].number_format = 'dd/mm/yyyy'
            except TypeError:
                pass
            c = Cleanser(str(test_proj_data[item.cell_key]))
            cleaned = c.clean()
            ws_res[item.cell_reference].value = cleaned
        elif item.template_sheet == config['TemplateSheets']['apm']:
            if has_whiff_of_total(item.cell_key):
                continue
            if isinstance(test_proj_data[item.cell_key], datetime.date):
                ws_apm[item.cell_reference].value = test_proj_data[item.cell_key]
                ws_apm[item.cell_reference].number_format = 'dd/mm/yyyy'
            if test_proj_data[item.cell_key] is None:
                continue
            try:
                if re.match(r'(\d+/\d+/\d+)', test_proj_data[item.cell_key]):
                    ws_apm[item.cell_reference].value = test_proj_data[item.cell_key]
                    ws_apm[item.cell_reference].number_format = 'dd/mm/yyyy'
            except TypeError:
                pass
            c = Cleanser(str(test_proj_data[item.cell_key]))
            cleaned = c.clean()
            ws_apm[item.cell_reference].value = cleaned
        elif item.template_sheet == config['TemplateSheets']['ap']:
            if isinstance(test_proj_data[item.cell_key], datetime.date):
                ws_ap[item.cell_reference].value = test_proj_data[item.cell_key]
                ws_ap[item.cell_reference].number_format = 'dd/mm/yyyy'
            if test_proj_data[item.cell_key] is None:
                continue
            try:
                if re.match(r'(\d+/\d+/\d+)', test_proj_data[item.cell_key]):
                    ws_ap[item.cell_reference].value = test_proj_data[item.cell_key]
                    ws_ap[item.cell_reference].number_format = 'dd/mm/yyyy'
            except TypeError:
                pass
            c = Cleanser(str(test_proj_data[item.cell_key]))
            cleaned = c.clean()
            ws_ap[item.cell_reference].value = cleaned
        elif item.template_sheet == config['TemplateSheets']['gmpp']:
            if isinstance(test_proj_data[item.cell_key], datetime.date):
                ws_gmpp[item.cell_reference].value = test_proj_data[item.cell_key]
                ws_gmpp[item.cell_reference].number_format = 'dd/mm/yyyy'
                continue
            try:
                if re.match(r'(\d+/\d+/\d+)', test_proj_data[item.cell_key]):
                    ws_gmpp[item.cell_reference].value = test_proj_data[item.cell_key]
                    ws_gmpp[item.cell_reference].number_format = 'dd/mm/yyyy'
            except TypeError:
                pass
            if test_proj_data[item.cell_key] is None:
                continue
            c = Cleanser(str(test_proj_data[item.cell_key]))
            cleaned = c.clean()
            ws_gmpp[item.cell_reference].value = cleaned

    imprint_current_quarter(ws_summary)
    lock_cells([
        ws_summary,
        ws_apm,
        ws_gmpp,
        ws_ap,
        ws_fb], TARGET_LOCK_CELLS)

    blank.save('/'.join([OUTPUT_DIR, '{}_{}_Return.xlsm'.format(
        test_proj.replace('/', '_'), config['QuarterData']['CurrentQuarter'])]))


def pop_all():
    """
    Populates the blank bicc_template file with data from the master, one
    form for each project dataset.
    """
    number_of_projects = len(get_list_projects(os.path.join(SOURCE_DIR, config['Master']['name'])))
    # we need to iterate through the master based on indexes so we use a range
    # based on the number of projects
    for p in range(number_of_projects):
        populate_blank_bicc_form(os.path.join(SOURCE_DIR, config['Master']['name']), p)


def get_dropdown_data(header=None):
    """
    Pull the dropdown data from the Dropdown List sheet in
    bicc_template.xlsx. Location of this template file might need
    to be dynamic.
    :return tuple of column values from sheet, with header value at list[0]:
    """
    wb = load_workbook(SOURCE_DIR + BLANK_TEMPLATE_FN, data_only=True)
    ws = wb['Dropdown List']
    columns = ws.columns
    col_lis = [col for col in columns]
    dropdown_data = [[c.value for c in t if c.value] for t in col_lis]
    if header:
        h = [h for h in dropdown_data if header in h[0]]
        h = tuple(h[0])
        # print("Getting {}".format(h))
        return h
    else:
        return dropdown_data


def get_dropdown_headers():
    wb = load_workbook(SOURCE_DIR + BLANK_TEMPLATE_FN)
    ws = wb['Dropdown']
    rows = ws.rows
    a_row = next(rows)
    return [h.value for h in a_row]


def create_validation(header):
    # if we need the regex to match the dropdown string - from pythex.org
    # dropdown_regex =
    # re.compile('"=\\'Dropdown List\\'!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)"')
    #

    try:
        f_str = VALIDATION_REFERENCES[header]
        dv = DataValidation(type='list', formula1=f_str, allow_blank=True)
        dv.prompt = "Please select from the list"
        dv.promptTitle = 'List Selection'
        return dv
    except KeyError:
        print("No validation")
        return


def main():
    parser = get_parser()
    args = vars(parser.parse_args())
    if args['loglevel']:
        log_lev = args['loglevel']
        logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler(OUTPUT_DIR + 'bcompiler.log', mode='w')
        fh.setLevel(logging.DEBUG)
        console = logging.StreamHandler()
        console.setLevel(log_lev)
        formatter = logging.Formatter('%(levelname)s - %(name)s - %(message)s')
        fh.setFormatter(formatter)
        console.setFormatter(colorlog.colorlog.ColoredFormatter())
        logger.addHandler(fh)
        logger.addHandler(console)
    else:
        logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler(OUTPUT_DIR + 'bcompiler.log', mode='w')
        fh.setLevel(logging.DEBUG)
        console = logging.StreamHandler()
        console.setLevel(logging.INFO)
        formatter = logging.Formatter('%(levelname)s - %(name)s - %(message)s')
        fh.setFormatter(formatter)
        console.setFormatter(colorlog.colorlog.ColoredFormatter())
        logger.addHandler(fh)
        logger.addHandler(console)

    if args['version']:
        print("{}".format(__version__))
        return
    if args['clean-datamap']:
        clean_datamap(DATAMAP_RETURN_TO_MASTER)
        print("datamap cleaned")
        return
    if args['transpose']:
        parse_csv_to_file(args['transpose'][0])
        return
    if args['populate']:
        master = '{}master.csv'.format(working_directory('source'))
        clean_datamap(DATAMAP_MASTER_TO_RETURN)
        parse_csv_to_file(master)
        populate_blank_bicc_form(master, args['populate'])
        return
    if args['all']:
        master = os.path.join(working_directory('source'), 'master.csv')
        clean_datamap(DATAMAP_RETURN_TO_MASTER)
        pop_all()
        return
    if args['analyser']:

        # checking for swimlane_milestones analyser
        if 'swimlane_milestones' in args['analyser']:
            analyser_args(args, swimlane_run)
            return

        # checking for swimlane_assurance_milestones analyser
        if 'swimlane_assurance_milestones' in args['analyser']:
            analyser_args(args, swimlane_assurance_run)
            return

        # checking for swimlane_milestones analyser
        if 'annex' in args['analyser']:
            analyser_args(args, annex_run)
            return

        if 'keyword' in args['analyser']:
            keyword_args(args, keyword_run)
            return

    if args['count-rows']:
        if args['csv'] and args['quiet']:
            logger.critical("-r option can only use --csv or --quiet, not both")
            return
        if args['csv']:
            row_data_formatter(csv_output=True)
        elif args['quiet']:
            row_data_formatter(quiet=True)
        else:
            row_data_formatter()
        return
    if args['compile'] and not args['compare']:
        if directory_has_returns_check(os.path.join(SOURCE_DIR, 'returns')):
            clean_datamap(DATAMAP_RETURN_TO_MASTER)
            compile_returns.run()
        else:
            sys.exit(1)
    if args['compare']:
        if directory_has_returns_check(os.path.join(SOURCE_DIR, 'returns')):
            clean_datamap(DATAMAP_RETURN_TO_MASTER)
            compile_returns.run(args['compare'])
        else:
            sys.exit(1)


if __name__ == '__main__':
    main()
