import logging
from openpyxl import load_workbook

logger = logging.getLogger('bcompiler.process.simple_comparitor')


class BCCell:

    def __init__(self, value, row_num=None, col_num=None, cellref=None):
        self.value = value
        self.row_num = row_num
        self.col_num = col_num
        self.cellref = cellref


class ParsedMaster:

    def __init__(self, master_file):
        self.master_file = master_file
        self._projects = []
        self._project_count = None
        self._key_col = []
        self._wb = load_workbook(self.master_file)
        self._ws = self._wb.active
        self._project_header_index = {}
        self._parse()

    def _parse(self):
        self._projects = [cell.value for cell in self._ws[1][1:]]
#       self._projects.sort()
        self._project_count = len(self.projects)
        self._key_col = [cell.value for cell in self._ws['A']]
        self._index_projects()

    @property
    def projects(self):
        return self._projects

    def _create_single_project_tuple(self, column=None, col_index=None):
        if col_index is None:
            col_data = self._ws[column]
            z = list(zip(self._key_col, col_data))
            return [((item[0]), (item[1].value)) for item in z]
        else:
            col_data = []
            for row in self._ws.iter_rows(
                min_row=1,
                max_col=col_index,
                min_col=col_index,
                max_row=len(self._key_col)
            ):
                count = 0
                for cell in row:
                    col_data.append(cell.value)
                    count += 1
            z = list(zip(self._key_col, col_data))
            return [((item[0]), (item[1])) for item in z]

    def _index_projects(self):
        self._project_header_index = {}
        for cell in self._ws[1]:
            if cell.value in self.projects:
                self._project_header_index[cell.value] = cell.col_idx

    def print_project_index(self):
        print('{:<68}{:>5}'.format("Project Title", "Column Index"))
        print('{:*^80}'.format(''))
        for k, v in self._project_header_index.items():
            print('{:<68}{:>5}'.format(k, v))

    def _create_dict_all_project_tuples(self):
        pass

    def __repr__(self):
        return "ParsedMaster for {}".format(
            self.master_file
        )

    def get_project_data(self, column=None, col_index=None):
        if column is None and col_index is None:
            raise TypeError('Please include at least one param')

        if column == 'A':
            raise TypeError("column must be 'B' or later in alphabet")

        if column:
            if isinstance(column, type('b')):
                data = self._create_single_project_tuple(column)
            else:
                raise TypeError('column must be a string')

        if col_index:
            if isinstance(col_index, type(1)):
                data = self._create_single_project_tuple(col_index=col_index)
            else:
                raise TypeError('col_index must be an integer')

        return data

    def _query_for_key(self, data, key):
        """
        Iterate through keys in output from get_project_data
        data list and return True if a key is found.
        """
        for item in data:
            if item[0] == key:
                self._query_result = item[1]
                return True

    def get_data_with_key(self, data, key):
        """
        Given a data list with project key/values in it (derived from
        a master spreadsheet, query a specific key to return a value.
        """
        # first query that the value exists
        if self._query_for_key(data, key):
            return self._query_result
        else:
            logger.info("No key {}".format(key))
            return None


def populate_cells(worksheet, bc_cells=[]):
    """
    Populate a worksheet with bc_cell object data.
    """
    for item in bc_cells:
        if item.cellref:
            worksheet[item.cellref].value = item.value
        else:
            worksheet.cell(
                row=item.row_num, column=item.col_num, value=item.value)
    return worksheet


class SimpleComparitor:
    """
    Simple method of comparing data in two master spreadsheets.
    """

    def __init__(self, masters=[]):
        """
        We want to get a list of master spreadsheets. These are simple
        file-references. The latest master should be master[-1].
        """
        if len(masters) > 2:
            raise ValueError("You can only analyse two spreadsheets.")

        self._masters = masters
        self._get_data()

    def _get_data(self):
        self._early_master = ParsedMaster(self._masters[0])
        self._current_master = ParsedMaster(self._masters[1])
        return (self._early_master, self._current_master)

    def compare(self, proj_index, key):
        project_data_early = self._early_master.get_project_data(
            col_index=proj_index)
        project_data_current = self._current_master.get_project_data(
            col_index=proj_index)
        return(
            self._early_master.get_data_with_key(project_data_early, key),
            self._current_master.get_data_with_key(project_data_current, key))
