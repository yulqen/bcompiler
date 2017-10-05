import configparser
import csv
import io
import os
import shutil
import tempfile
from datetime import datetime

import pytest
from openpyxl import load_workbook, Workbook

from ..utils import generate_test_template_from_real as gen_template

TEMPDIR = tempfile.gettempdir()

AUX_DIR = "/".join([TEMPDIR, 'bcompiler'])
SOURCE_DIR = "/".join([AUX_DIR, 'source'])
RETURNS_DIR = "/".join([SOURCE_DIR, 'returns'])
OUTPUT_DIR = "/".join([AUX_DIR, 'output'])

try:
    os.mkdir(AUX_DIR)
except (FileExistsError, IsADirectoryError):
    shutil.rmtree(AUX_DIR)
    os.mkdir(AUX_DIR)
    os.mkdir(OUTPUT_DIR)
    os.mkdir(SOURCE_DIR)
    os.mkdir(RETURNS_DIR)

config = configparser.ConfigParser()
CONFIG_FILE = 'test_config.ini'
config.read(CONFIG_FILE)

BICC_TEMPLATE_FOR_TESTS = config['Template']['ActualTemplatePath']

datamap_header = "cell_key,template_sheet,cell_reference,verification"

datamap_data = """
Project/Programme Name,Summary,B5,
SRO Sign-Off,Summary,B49,
Reporting period (GMPP - Snapshot Date),Summary,G3,
Quarter Joined,Summary,I3,
GMPP (GMPP - formally joined GMPP),Summary,G5,
IUK top 40,Summary,G6,
Top 37,Summary,I5,
DfT Business Plan,Summary,I6,
DFT ID Number,Summary,B6,
MPA ID Number,Summary,C6,
Working Contact Name,Summary,H8,
Working Contact Telephone,Summary,H9,
SRO Tenure Start Date,Summary,C15,
SRO Tenure End Date,Summary,C17,
Working Contact Email,Summary,H10,
DfT Group,Summary,B8,DfT Group,
DfT Division,Summary,B9,DfT Division,
Agency or delivery partner (GMPP - Delivery Organisation primary),Summary,B10,Agency,
Strategic Alignment/Government Policy (GMPP - Key drivers),Summary,B26,
Project stage,Approval & Project milestones,B5,Project stage,
Project stage if Other,Approval & Project milestones,D5,
Last time at BICC,Approval & Project milestones,B4,
Next at BICC,Approval & Project milestones,D4,
Approval MM1,Approval & Project milestones,A9,
Approval MM1 Original Baseline,Approval & Project milestones,B9,
Approval MM1 Latest Approved Baseline,Approval & Project milestones,C9,
Approval MM1 Forecast / Actual,Approval & Project milestones,D9,
Approval MM1 Milestone Type,Approval & Project milestones,E9,Milestone Types,
Approval MM1 Notes,Approval & Project milestones,F9,
Approval MM2,Approval & Project milestones,A10,
Approval MM2 Original Baseline,Approval & Project milestones,B10,
Approval MM2 Latest Approved Baseline,Approval & Project milestones,C10,
Approval MM2 Forecast / Actual,Approval & Project milestones,D10,
Approval MM2 Milestone Type,Approval & Project milestones,E10,
Approval MM2 Notes,Approval & Project milestones,F10,
Approval MM3,Approval & Project milestones,A11,
Approval MM3 Original Baseline,Approval & Project milestones,B11,
Approval MM3 Latest Approved Baseline,Approval & Project milestones,C11,
Approval MM3 Forecast / Actual,Approval & Project milestones,D11,
Approval MM3 Milestone Type,Approval & Project milestones,E11,Milestone Types,
Approval MM3 Notes,Approval & Project milestones,F11,
Significant Steel Requirement,Finance & Benefits,D15,Yes/No,
SRO Finance confidence,Finance & Benefits,C6,RAG 2,
BICC approval point,Finance & Benefits,E9,Business Cases,
Latest Treasury Approval Point (TAP) or equivalent,Finance & Benefits,E10,Business Cases,
Business Case used to source figures (GMPP TAP used to source figures),Finance & Benefits,C9,Business Cases,
Date of TAP used to source figures,Finance & Benefits,E11,
Name of source in not Business Case (GMPP -If not TAP please specify equivalent document used),Finance & Benefits,C10,
If not TAP please specify date of equivalent document,Finance & Benefits,C11,
Version Number Of Document used to Source Figures (GMPP - TAP version Number),Finance & Benefits,C12,
Date document approved by SRO,Finance & Benefits,C13,
Real or Nominal - Baseline,Finance & Benefits,C18,Finance figures format,
Real or Nominal - Actual/Forecast,Finance & Benefits,E18,Finance figures format,
Index Year,Finance & Benefits,B19,Index Years,
Deflator,Finance & Benefits,B20,Finance type,
Source of Finance,Finance & Benefits,B21,Finance type,
Other Finance type Description,Finance & Benefits,D21,
NPV for all projects and NPV for programmes if available,Finance & Benefits,B22,
Project cost to closure,Finance & Benefits,B23,
RDEL Total Budget/BL,Finance & Benefits,C72,
CDEL Total Budget/BL,Finance & Benefits,C125,
Non-Gov Total Budget/BL,Finance & Benefits,C135,
Total Budget/BL,Finance & Benefits,C136,
RDEL Total Forecast,Finance & Benefits,D133,
CDEL Total Forecast,Finance & Benefits,D134,
Non-Gov Total Forecast,Finance & Benefits,D135,
Total Forecast,Finance & Benefits,D136,
RDEL Total Variance,Finance & Benefits,E133,
CDEL Total Variance,Finance & Benefits,E134,
Assurance MM1,Assurance Planning,A8,
Assurance MM1 Original Baseline,Assurance Planning,B8,
Assurance MM1 Latest Approved Baseline,Assurance Planning,C8,
Assurance MM1 Forecast - Actual,Assurance Planning,D8,
Assurance MM1 DCA,Assurance Planning,E8,RAG,
Assurance MM1 Notes,Assurance Planning,F8,
Assurance MM2,Assurance Planning,A9,
Assurance MM2 Original Baseline,Assurance Planning,B9,
Assurance MM2 Latest Approved Baseline,Assurance Planning,C9,
Assurance MM2 Forecast - Actual,Assurance Planning,D9,
Assurance MM2 DCA,Assurance Planning,E9,RAG,
Assurance MM2 Notes,Assurance Planning,F9,
Total Number of public sector employees working on the project,Resource,C37,
Total Number of external contractors working on the project,Resource,E37,
Total Number or vacancies on the project,Resource,G37,
Resources commentary,Resource,C19,
Total number of employees funded to work on project,Resource,I17,
Resources commentary,Resource,C19,
Overall Resource DCA - Now,Resource,I38,Capability RAG,
Overall Resource DCA - Future,Resource,J38,Capability RAG,
Digital - Now,Resource,I25,Capability RAG,
Digital - Future,Resource,J25,Capability RAG,
Information Technology - Now,Resource,I26,Capability RAG,
Information Technology - Future,Resource,J26,Capability RAG,
Legal Commercial Contract Management - Now,Resource,I27,Capability RAG,
Legal Commercial Contract Management - Future,Resource,J27,Capability RAG,
Project Delivery - Now,Resource,I28,Capability RAG,
Project Delivery - Future,Resource,J28,Capability RAG,
Change Implementation - Now,Resource,I29,Capability RAG,
Change Implementation - Future,Resource,J29,Capability RAG,
Technical - Now,Resource,I30,Capability RAG,
Technical - Future,Resource,J30,Capability RAG,
Industry Knowledge - Now,Resource,I31,Capability RAG,
Industry Knowledge - Future,Resource,J31,Capability RAG,
Finance - Now,Resource,I32,Capability RAG,
Finance - Future,Resource,J32,Capability RAG,
Analysis Now,Resource,I33,Capability RAG,
Analysis - future,Resource,J33,Capability RAG,
"""


@pytest.fixture(scope='module')
def blank_template():
    gen_template(BICC_TEMPLATE_FOR_TESTS, SOURCE_DIR)
    output_file = '/'.join([SOURCE_DIR, 'gen_bicc_template.xlsm'])
#   yield output_file
    return output_file
#   os.remove(output_file)


@pytest.fixture(scope='module')
def datamap():
    name = 'datamap.csv'
    s = io.StringIO()
    s.write(datamap_header)
    s.write(datamap_data)
    s.seek(0)
    s_string = s.readlines()
#   del s_string[0]
    with open('/'.join([SOURCE_DIR, name]), 'w') as csv_file:
        for x in s_string:
            csv_file.write(x)
    return '/'.join([SOURCE_DIR, name])


@pytest.fixture(scope='module')
def populated_template():
    gen_template(BICC_TEMPLATE_FOR_TESTS, SOURCE_DIR)
    datamap()
    dm = "/".join([SOURCE_DIR, 'datamap.csv'])
    wb = load_workbook("/".join([SOURCE_DIR, 'gen_bicc_template.xlsm']), keep_vba=True)
    output_file = "/".join([RETURNS_DIR, 'populated_test_template.xlsm'])
    for fl in range(10):
        with open(dm, 'r', newline='') as f:
            reader = csv.DictReader(f)
            for line in reader:
                if line['cell_key'].startswith('Date'):  # we want to test date strings
                    wb[line['template_sheet']][line['cell_reference']].value = "20/06/2017"
                elif line['cell_key'].startswith('SRO Tenure'):  # we want to test date strings
                    wb[line['template_sheet']][line['cell_reference']].value = "10/08/2017"
                else:
                    wb[line['template_sheet']][line['cell_reference']].value = " ".join([line['cell_key'].upper(), str(fl)])
            output_file = "/".join([RETURNS_DIR, 'populated_test_template{}.xlsm'
                                    .format(fl)])
            wb.save(output_file)
    # we save 10 of them but only return the first for testing
    return output_file


def split_datamap_line(line: tuple):
    for item in line:
        yield item


@pytest.fixture(scope='module')
def master():
    """
    This is master file created for the purpose of using a base for bcompiler -a, which
    populates all the returns. It simply takes the field name from the datamap and
    puts it in upper case and appends a digit (1, 2 or 3 because we're only simulating
    a master with 3 projects here.
    :return: output_file
    """
    wb = Workbook()
    output_file = "/".join([OUTPUT_DIR, 'master.xlsx'])
    ws = wb.active
    ws.title = "Master for Testing"
    for item in enumerate(datamap_data.split('\n')):
        if not item[0] == 0 and not item[1] == "":
            g = split_datamap_line(item)
            next(g)
            ix = next(g).split(',')[0]
            ws[f"A{str(item[0])}"] = ix
            if item[1].startswith('Date'):  # testing for date objects
                ws[f"B{str(item[0])}"] = datetime(2017, 6, 20)
                ws[f"C{str(item[0])}"] = datetime(2017, 6, 20)
                ws[f"D{str(item[0])}"] = datetime(2017, 6, 20)
            elif item[1].startswith('SRO Tenure'):  # testing for date objects
                ws[f"B{str(item[0])}"] = datetime(2017, 8, 10)
                ws[f"C{str(item[0])}"] = datetime(2017, 8, 10)
                ws[f"D{str(item[0])}"] = datetime(2017, 8, 10)
            else:
                ws[f"B{str(item[0])}"] = " ".join([ix.upper(), "1"])
                ws[f"C{str(item[0])}"] = " ".join([ix.upper(), "2"])
                ws[f"D{str(item[0])}"] = " ".join([ix.upper(), "3"])
    wb.save(output_file)
    return output_file

@pytest.fixture(scope='module')
def previous_quarter_master():
    """
    This is a replica of the master() fixture above, but we're changing some
    values to simulate an earlier master than needs to be compared against.

    The values we're amending are the three values for "Working Contact Name",
    which appear in cells B11, C11, D11.
    :return: output_file
    """
    wb = Workbook()
    output_file = "/".join([OUTPUT_DIR, 'early_master.xlsx'])
    ws = wb.active
    ws.title = "Previous quarter master for testing"
    for item in enumerate(datamap_data.split('\n')):
        if not item[0] == 0 and not item[1] == "":
            g = split_datamap_line(item)
            next(g)
            ix = next(g).split(',')[0]
            ws[f"A{str(item[0])}"] = ix
            if item[1].startswith('Date'):  # testing for date objects
                ws[f"B{str(item[0])}"] = datetime(2017, 6, 20)
                ws[f"C{str(item[0])}"] = datetime(2017, 6, 20)
                ws[f"D{str(item[0])}"] = datetime(2017, 6, 20)
            elif item[1].startswith('SRO Tenure'):  # testing for date objects
                ws[f"B{str(item[0])}"] = datetime(2017, 8, 10)
                ws[f"C{str(item[0])}"] = datetime(2017, 8, 10)
                ws[f"D{str(item[0])}"] = datetime(2017, 8, 10)
            else:
                ws[f"B{str(item[0])}"] = " ".join([ix.upper(), "1"])
                ws[f"C{str(item[0])}"] = " ".join([ix.upper(), "2"])
                ws[f"D{str(item[0])}"] = " ".join([ix.upper(), "3"])
    # here we amend the three string cells...
    ws['B11'].value = ' '.join([ws['B11'].value, 'AMENDED'])
    ws['C11'].value = ' '.join([ws['B11'].value, 'AMENDED'])
    ws['D11'].value = ' '.join([ws['B11'].value, 'AMENDED'])

    # here we amend a single date cells...
    # this is for "SRO Tenure Start Date"
    ws['B13'].value = datetime(2017, 3, 1)

    # now setting an later date for "SRO Tenure End Date"
    # also now for PROJECT/PROGRAMME NAME 2
    ws['C14'].value = datetime(2019, 6, 6)

    wb.save(output_file)
    return output_file
