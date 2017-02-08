import os
import random
from datetime import datetime
import pytest

from bcompiler.process.digest import Series

from openpyxl import Workbook

ws_summary_B5_rand = [
    'Cookfield Rebuild',
    'Smithson Glenn Park Editing',
    'Brinkles Bypass Havensmere',
    'Folicles On Fire Ltd Extradition',
    'Puddlestein Havelock Underpass',
]

ws_summary_B8_rand = [
    'Aerobics, Maritime and Commerce',
    'TSAD',
    'Special Transport Needs for the Northern Populace',
    'Parenting, Levels and Financial Irregularity',
    'HR',
]


ws_finance_C6_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]

ws_finance_C11_rand = [
    datetime(2011, 1, 23),
    datetime(2012, 2, 22),
    datetime(2010, 3, 13),
    datetime(2018, 8, 1),
    datetime(2001, 11, 12),
]


ws_finance_B19_rand = [
    '2012',
    '2013',
    '2011',
    '2018',
    '2002',
    '2007',
]

ws_finance_C18_rand = [
    'Real',
    'Nominal',
]

ws_finance_C36_rand = [
    2.00,
    4.20,
    1.13,
    12.09,
    222.07,
]

ws_finance_C44_rand = [
    12.00,
    41.20,
    13.13,
    122.09,
    22.07,
]

ws_finance_C77_rand = [
    29.00,
    49.23,
    23.43,
    1.89,
    290.37,
]


ws_resources_C7_rand = [
    9.00,
    19.00,
    29.5,
    12.00,
    20.5,
]

ws_resources_G17_rand = [
    9.90,
    19.22,
    29.93,
    1202.89,
    20.37,
]


ws_resources_I30_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]


ws_resources_J30_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]


ws_resources_J38_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]

ws_approval_C10_rand = [
    datetime(2001, 2, 23),
    datetime(2013, 7, 22),
    datetime(2013, 2, 13),
    datetime(2013, 4, 1),
    datetime(2011, 1, 12),
]


ws_approval_F19_rand = [
    'A load of absolute horseradish.',
    'When people speak of these kind of things, they are often surprised.',
    'It is very bad here. Completely unacceptable when you think about it.',
    'Never worry too much about it - it wont last forever',
    'There is a forester on this project who is disrupting everything.'
]


ws_approval_B39_rand = [
    datetime(2101, 2, 23),
    datetime(2023, 7, 22),
    datetime(2023, 2, 13),
    datetime(2019, 5, 1),
    datetime(2021, 1, 12),
]


ws_assurance_C4_rand = [
    datetime(2012, 2, 20),
    datetime(2013, 7, 12),
    datetime(2003, 2, 19),
    datetime(2017, 5, 10),
    datetime(2016, 6, 18),
]


ws_assurance_D10_rand = [
    datetime(2002, 2, 23),
    datetime(2003, 3, 2),
    datetime(2013, 6, 9),
    datetime(2013, 9, 12),
    datetime(2017, 12, 17),
]


ws_resources_E17_rand = [
    'Green',
    'Amber/Green',
    'Amber',
    'Amber/Red',
    'Red',
]


@pytest.fixture
def bicc_return():
    wb = Workbook()
    wb.create_sheet('Summary')
    wb.create_sheet('Finance & Benefits')
    wb.create_sheet('Approval & Project milestones')
    wb.create_sheet('Resources')
    wb.create_sheet('Assurance planning')
    wb.create_sheet('GMPP info')

    # Summary fixture
    ws_summary = wb['Summary']
    ws_summary['A5'].value = 'Project/Programme Name'
    ws_summary['B5'].value = random.choice(ws_summary_B5_rand)
    ws_summary['A8'].value = 'DfT Group'
    ws_summary['B8'].value = random.choice(ws_summary_B8_rand)

    # Finance & Benefits fixture
    ws_finance = wb['Finance & Benefits']
    ws_finance['A6'].value = 'SRO Finance Confidence'
    ws_finance['C6'].value = random.choice(ws_finance_C6_rand)
    ws_finance['B11'].value = 'Date of Business Case'
    ws_finance['C11'].value = random.choice(ws_finance_C11_rand)
    ws_finance['A19'].value = 'Index Year'
    ws_finance['B19'].value = random.choice(ws_finance_B19_rand)
    ws_finance['A18'].value = 'Real or Nominal'
    ws_finance['C18'].value = random.choice(ws_finance_C18_rand)
    ws_finance['A36'].value = '2019/2020'
    ws_finance['C36'].value = random.choice(ws_finance_C36_rand)
    ws_finance['A44'].value = 'Total'
    ws_finance['C44'].value = random.choice(ws_finance_C44_rand)
    ws_finance['A77'].value = 'Total WLC (RDEL)'
    ws_finance['C77'].value = random.choice(ws_finance_C77_rand)

    # Resources fixture
    ws_resources = wb['Resources']
    ws_resources['A7'].value = 'SCS(PB2)'
    ws_resources['C7'].value = random.choice(ws_resources_C7_rand)
    ws_resources['A17'].value = 'Total'
    ws_resources['G17'].value = random.choice(ws_resources_G17_rand)
    ws_resources['A30'].value = 'Change Implementation'
    ws_resources['I30'].value = random.choice(ws_resources_I30_rand)
    ws_resources['J30'].value = random.choice(ws_resources_J30_rand)
    ws_resources['G38'].value = 'Overall Assessment'
    ws_resources['J38'].value = random.choice(ws_resources_J38_rand)

    # Approval and Project Milestones fixture
    ws_approvals = wb['Approval & Project milestones']
    ws_approvals['A10'].value = 'SOBC - HMT Approval'
    ws_approvals['C10'].value = random.choice(ws_approval_C10_rand)
    ws_approvals['A19'].value = 'FBC - HMT Approval'
    ws_approvals['F19'].value = random.choice(ws_approval_F19_rand)
    ws_approvals['A39'].value = 'Completion of Construction'
    ws_approvals['B39'].value = random.choice(ws_approval_B39_rand)

    # Assurance fixture
    ws_assurance = wb['Assurance planning']
    ws_assurance['B4'].value = 'Date Created'
    ws_assurance['C4'].value = random.choice(ws_assurance_C4_rand)
    ws_assurance['A10'].value = 'Gate 0 (Programme)'
    ws_assurance['D10'].value = random.choice(ws_assurance_D10_rand)
    ws_assurance['A17'].value = 'Review Point 4 MPRG'
    ws_assurance['E17'].value = 'Amber/Green'

    wb.save('/tmp/test-bicc-return.xlsx')
    yield '/tmp/test-bicc-return.xlsx'
    os.unlink('/tmp/test-bicc-return.xlsx')


@pytest.fixture
def series():
    series = Series('Financial Quarters')
    return series
