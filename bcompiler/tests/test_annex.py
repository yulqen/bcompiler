import datetime

from openpyxl import load_workbook

from ..analysers.annex import run as annex_run, abbreviate_project_stage
from ..analysers.annex import _dca_map


def test_annex_title(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 2_ANNEX.xlsx'))
    ws = wb.active
    assert ws['A1'].value == 'PROJECT/PROGRAMME NAME 2'


def test_pound_sign(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 1_ANNEX.xlsx'))
    ws = wb.active
    assert ws['A5'].value == 'WLC(£m):'


def test_b5_one_decimal(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 1_ANNEX.xlsx'))
    ws = wb.active
    assert ws['B5'].value == '32.3'


def test_abbr_func():
    assert abbreviate_project_stage('Outline Business Case') == 'OBC'
    assert abbreviate_project_stage('Strategic Outline Case') == 'SOBC'
    assert abbreviate_project_stage('Strategic Outline Business Case') == 'SOBC'
    assert abbreviate_project_stage('Full Business Case') == 'FBC'
    assert abbreviate_project_stage('General Turgidson') == 'UNKNOWN STAGE'


def test_abbreviate_stage_name(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 1_ANNEX.xlsx'))
    ws = wb.active
    assert ws['D5'].value == 'SOBC'


def test_correct_date_format(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 1_ANNEX.xlsx'))
    ws = wb.active
    assert isinstance(ws['F5'].value, datetime.date)


def test_amber_green_cell_colour(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 1_ANNEX.xlsx'))
    ws = wb.active
    assert ws['B7'].fill.fgColor.rgb == '00f9cb31'


def test_dca_map(previous_quarter_master, tmpdir, master):
    annex_run(previous_quarter_master, [str(tmpdir)], master)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME NAME 1_ANNEX.xlsx'))
    wb.save('/home/lemon/Desktop/tits.xlsx')
    ws = wb.active
    assert ws['D7'].value == 'Amber'
