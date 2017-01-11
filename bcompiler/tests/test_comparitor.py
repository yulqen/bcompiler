import pytest

from bcompiler.process.simple_comparitor import BCCell, ParsedMaster
from bcompiler.process.simple_comparitor import FileComparitor
from bcompiler.process.simple_comparitor import populate_cells

from bcompiler.utils import cell_bg_colour

from openpyxl import Workbook

key_col_data = [
    'Project/Programme Name',
    'SRO Sign-Off',
    'Reporting period (GMPP - Snapshot Date)',
    'Quarter Joined',
    'GMPP (GMPP – formally joined GMPP)',
    'IUK top 40',
    'Top 37',
    'DfT Business Plan',
    'GMPP - IPA ID Number',
    'DFT ID Number',
    'Working Contact Name',
    'Working Contact Telephone',
    'Working Contact Email',
    'DfT Group',
    'DfT Division',
    'Agency or delivery partner (GMPP - Delivery Organisation primary)',
    'Strategic Alignment/Government Policy (GMPP – Key drivers)',
    'Project Scope',
    'Brief project description (GMPP – brief descripton)',
    'Delivery Structure',
    'Description if \'Other',
    'Change Delivery Methodology'
]

project_b_data = [
    'Digital Signalling',
    '2016-10-12 0:00:00',
    'Q2 1617',
    None,
    None,
    None,
    None,
    None,
    None,
    8,
    'Niall Le Mage',
    '2079442043',
    'niall.lemage@dft.gsi.gov.uk',
    'Rail Group',
    'Network Services',
    'Network Rail',
    'In line with DfTs single Departmental Plan to roll out new technology',
    'Scope of the ETCS cab-fitment fund: | to facilitate the inclusion of',
    'The fitting of digital signalling technology to prototype passenger',
    'Project',
    None,
    'Waterfall',
]


@pytest.fixture
def populate_test_data():
    wb = Workbook()
    ws = wb.active

    for item in key_col_data:
        c = BCCell(item, row_num=key_col_data.index(item) + 1, col_num=1)
        ws.cell(value=c.value, row=c.row_num, column=c.col_num)
    ws['C1'].fill = cell_bg_colour([255, 0, 0])
    yield ws


@pytest.fixture
def parsed_master():
    pm = ParsedMaster(populate_test_data)
    return pm


def test_populate_test_data(populate_test_data):
    assert populate_test_data['A1'].value == 'Project/Programme Name'
    assert populate_test_data['A2'].value == 'SRO Sign-Off'


def test_populate_function():
    wb = Workbook()
    ws = wb.active

    # populate by coordinates
    populate_cells(ws, [BCCell("Test1", cellref="A1")])
    assert ws['A1'].value == "Test1"

    # populate by row, col
    populate_cells(ws, [BCCell("Fanciso Monk", 2, 1)])
    assert ws['A2'].value == "Fanciso Monk"


def test_cell_colours(populate_test_data):
    assert populate_test_data['C1'].fill.fgColor.rgb == '00FF0000'


SOURCE_EARLY = ('/home/lemon/Documents/bcompiler/'
                'output/compiled_master_early.xlsx')
SOURCE_CURRENT = ('/home/lemon/Documents/bcompiler/'
                  'output/compiled_master_current.xlsx')


def test_parsed_master():
    pm = ParsedMaster(SOURCE_CURRENT)
    project_data = pm.get_project_data(col_index=2)
    project_data2 = pm.get_project_data(col_index=13)
    assert 'Crossrail Programme' in pm.projects
    assert pm.get_project_data(
        'C')[0][1] == 'Search and Rescue Helicopters'
    assert pm.get_project_data(
        col_index=3)[0][1] == 'Search and Rescue Helicopters'
    assert pm.get_data_with_key(project_data, 'Top 37') is None
    assert pm.get_data_with_key(
        project_data2, 'Project/Programme Name') == 'DfT Headquarters'


def test_comparitor():
    comp = FileComparitor([SOURCE_EARLY, SOURCE_CURRENT])
    t = comp.compare(2, 'Project/Programme Name')
    t2 = comp.compare(3, 'Project/Programme Name')
    assert t[0] == 'Digital Signalling'
    assert t2[0] == 'Search and Rescue Helicopters'


def test_single_comparitor():
    comp = FileComparitor([SOURCE_EARLY])
    t = comp.compare(2, 'Project/Programme Name')
    t2 = comp.compare(3, 'Project/Programme Name')
    t3 = comp.compare(5, 'Working Contact Name')
    assert t == 'Digital Signalling'
    assert t2 == 'Search and Rescue Helicopters'
    assert t3 == 'Jonathan Daley'
