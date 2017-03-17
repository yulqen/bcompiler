import pytest
import os
from openpyxl import Workbook

import bcompiler.tests.fixtures

BICC_RETURN_FIXTURE = bcompiler.tests.fixtures.bicc_return


@pytest.fixture
def bicc_return():
    wb = Workbook()
    wb.create_sheet('Summary')
    wb.create_sheet('Approval & Project milestones')
    wb.create_sheet('Finance & Benefits')
    wb.create_sheet('Resources')
    wb.create_sheet('Assurance planning')
    wb.create_sheet('GMPP info')
    ws = wb['Summary']
    # enter some values in the right slots
    ws['B5'].value = 'Cookfield Rebuild'
    ws['B8'].value = 'Roads, Monitoring and Horse'

    wb.save('/tmp/test-bicc-return.xlsx')
    yield '/tmp/test-bicc-return.xlsx'
    os.unlink('/tmp/test-bicc-return.xlsx')
