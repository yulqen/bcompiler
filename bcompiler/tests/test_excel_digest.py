from pymongo import MongoClient
import pytest

from datetime import date


@pytest.fixture(scope="module")
def mock_datamap_object():
    """
    Mock a Datamap object to use for data retrieval, in MongoDB.

    Datamap object has the format:
        type: <BICC/GMPP/Etc>
        date: <date>
        version: <version>
        project_programme:
            { cellref: <cellref>,
              sheet: <sheetname>,
              verification_list:
                  { name: <list_name>,
                    list_items: [<item 1>, <item 2>,...] }
              conditional_formatting:
                  { rule1: <rule>,
                    rule2: <rule>,
                    ... }
              etc ...}

    The structure above can be used to create a Cell object which is used when
    populating the sheet.
    """
    client = MongoClient()
    db = client.test_database
    collection = db.test_collection
    test_data = {'Map Type': 'BICC',
                 'First Name': 'B1',
                 'Last Name': 'B2',
                 'Age': 'B3',
                 'Birthday': 'B4',
                 }
    collection.insert_one(test_data)
    yield collection
    print("Teardown MongoDB test_database")
    client.drop_database(db)


@pytest.fixture
def mock_bicc_form():
    """Set up a basic form for testing."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "First Name"
    ws['B1'] = "Matthew"
    ws['A2'] = "Last Name"
    ws['B2'] = "Lemon"
    ws['A3'] = "Age"
    ws['B3'] = 41
    ws['A4'] = "Birthday"
    ws['B4'] = date(1975, 1, 24)


def test_db(mock_datamap_object):
    doc = list(mock_datamap_object.find({"Map Type": {"$eq": "BICC"}}))
    d = {}
    assert type(doc[0]) == type(d)
    assert doc[0]['Map Type'] == 'BICC'
    assert doc[0]['First Name'] == 'B1'


def test_populate_datamap_from_csv():
    pass


def test_populate_datamap_from_xlsx():
    pass


def test_for_existence_of_datamaps():
    pass
