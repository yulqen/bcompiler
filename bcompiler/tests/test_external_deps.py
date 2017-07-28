"""
This module tests external dependencies in the bcompiler system,
including the datamap and the bicc_template.xlsx.

If the template changes, the cell references in test_generated_template()
function below should be ameneded in the test_config.ini file. All cellrefs
in this file are fixed - their contents may change with template changes.
"""
import configparser
import csv

from openpyxl import load_workbook

import bcompiler.compile as compile_module
from ..compile import parse_source_cells as parse, run

config = configparser.ConfigParser()
CONFIG_FILE = 'test_config.ini'
config.read(CONFIG_FILE)


def test_existence(datamap):
    with open(datamap, 'r', newline='') as f:
        assert next(f).startswith('Project/Programme Name')
        reader = csv.reader(f)
        assert next(reader)[2] == 'B49'


def test_generated_template(blank_template):
    wb = load_workbook(blank_template)
    sheet_s = wb[config['TemplateTestData']['summary_sheet']]
    sheet_fb = wb[config['TemplateTestData']['fb_sheet']]
    sheet_r = wb[config['TemplateTestData']['resource']]
    sheet_apm = wb[config['TemplateTestData']['apm']]
    sheet_ap = wb[config['TemplateTestData']['ap']]
    assert sheet_s['A8'].value == config['SummaryData']['A8']
    assert sheet_s['A46'].value == config['SummaryData']['A46']
    assert sheet_fb['A121'].value == config['FinanceData']['A121']
    assert sheet_fb['F26'].value == config['FinanceData']['F26']
    assert sheet_r['A36'].value == config['ResourceData']['A36']
    assert sheet_r['A12'].value == config['ResourceData']['A12']
    assert sheet_apm['E7'].value == config['ApprovalProjectMilestones']['E7']
    assert sheet_ap['B32'].value == config['AssurancePlanning']['B32']
    assert sheet_ap['C31'].value == None  # can't put None value in config file


def test_incorrect_template_cells(blank_template):
    wb = load_workbook(blank_template)
    sheet_apm = wb['Approval & Project milestones']
    assert sheet_apm['A43'].value == None
    assert sheet_apm['A430'].value == None


# the test data is just the field name uppercased
# check the fixture code if you don't believe me
def test_populated_template(populated_template):
    wb = load_workbook(populated_template)
    sheet_summary = wb['Summary']
    sheet_fb = wb['Finance & Benefits']
    sheet_r = wb['Resource']
    sheet_apm = wb['Approval & Project milestones']
    sheet_ap = wb['Assurance Planning']
    assert sheet_summary['B5'].value == "PROJECT/PROGRAMME NAME"
    assert sheet_summary['B10'].value == "AGENCY OR DELIVERY PARTNER (GMPP - DELIVERY ORGANISATION PRIMARY)"
    assert sheet_summary['H10'].value == "WORKING CONTACT EMAIL"
    assert sheet_fb['C18'].value == "REAL OR NOMINAL - BASELINE"
    assert sheet_r['I25'].value == 'DIGITAL - NOW'
    assert sheet_apm['B9'].value == 'APPROVAL MM1 ORIGINAL BASELINE'
    assert sheet_ap['D8'].value == 'ASSURANCE MM1 FORECAST - ACTUAL'


def test_compile(populated_template, datamap):
    data = parse(populated_template, datamap)
    assert data[0]['gmpp_key'] == 'Project/Programme Name'
    assert data[0]['gmpp_key_value'] == 'PROJECT/PROGRAMME NAME'


def test_run(datamap):
    # print([item for item in dir(compile_module) if not item.startswith("__")])
    # patching module attributes to get it working
    setattr(compile_module, 'RETURNS_DIR', '/tmp/bcompiler-test/')
    setattr(compile_module, 'OUTPUT_DIR', '/tmp/bcompiler-test-output/')
    setattr(compile_module, 'TODAY',  '2017-07-28')
    setattr(compile_module, 'DATAMAP_RETURN_TO_MASTER', datamap)
    run()

