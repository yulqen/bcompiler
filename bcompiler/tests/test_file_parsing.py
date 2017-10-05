import os
import tempfile
from datetime import date

from openpyxl import load_workbook

import bcompiler.compile as compile_module
from ..compile import parse_source_cells as parse
from ..compile import run
from ..process.datamap import Datamap

TODAY = date.today().isoformat()
TEMPDIR = tempfile.gettempdir()

AUX_DIR = "/".join([TEMPDIR, 'bcompiler'])
SOURCE_DIR = "/".join([AUX_DIR, 'source'])
RETURNS_DIR = "/".join([SOURCE_DIR, 'returns'])
OUTPUT_DIR = "/".join([AUX_DIR, 'output'])


def test_populate_single_template_from_master(populated_template, datamap):
    """
    This tests bcompiler -b X essentially (or bcompiler -a, but for a single file.
    """
    data = parse(populated_template, datamap)
    assert data[0]['gmpp_key'] == 'Project/Programme Name'
    assert data[0]['gmpp_key_value'] == 'PROJECT/PROGRAMME NAME 9'


def test_compile_all_returns_to_master_no_comparison(populated_template, datamap):
    """
    This tests 'bcompiler compile' or 'bcompiler' option.
    """
    # print([item for item in dir(compile_module) if not item.startswith("__")])
    # patching module attributes to get it working
    setattr(compile_module, 'RETURNS_DIR', RETURNS_DIR)
    setattr(compile_module, 'OUTPUT_DIR', OUTPUT_DIR)
    setattr(compile_module, 'TODAY', date.today().isoformat())
    setattr(compile_module, 'DATAMAP_RETURN_TO_MASTER', datamap)
    run()
    # for one of the templates that we have compiled (using 9 I think)...
    # get the project title...
    data = parse(populated_template, datamap)
    project_title = data[0]['gmpp_key_value']
    # then we need to open up the master that was produced by run() function above...
    wb = load_workbook(os.path.join(OUTPUT_DIR, 'compiled_master_{}_{}.xlsx'.format(TODAY, "Q2")))
    ws = wb.active
    # we then need the "Project/Programme Name" row from the master
    project_title_row = [i.value for i in ws[1]]
    assert project_title in project_title_row




def test_datamap_class(datamap):
    """
    This tests correct creation of Datamap object.
    """
    dm = Datamap()
    dm.cell_map_from_csv(datamap)
    assert dm.cell_map[1].cell_reference == 'B49'
