import pytest

from openpyxl import load_workbook

from ..analysers.financial import run as financial_run
from ..analysers.utils import project_titles_in_master


def test_a1_cell(tmpdir, master):
    tmpdir = [tmpdir]
    financial_run([master], output_path=tmpdir)
    tmpdir = tmpdir[0]
    wb = load_workbook(tmpdir.join('financial_analysis.xlsx'))
    ws = wb.get_sheet_by_name('PROJECT_PROGRAMME NAME 1')
    assert ws['A1'].value == 'PROJECT/PROGRAMME NAME 1'


def test_header_cells(tmpdir, master):
    tmpdir = [tmpdir]
    financial_run([master], output_path=tmpdir)
    tmpdir = tmpdir[0]
    wb = load_workbook(tmpdir.join('financial_analysis.xlsx'))
    ws = wb.get_sheet_by_name('PROJECT_PROGRAMME NAME 1')
    assert ws['A3'].value == 'Q1 18/19'
