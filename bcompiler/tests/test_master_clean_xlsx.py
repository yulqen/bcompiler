"""
After data from the BICC forms is migrated into a compiled_master.xlsx file,
we need to avoid having to re-save this file as csv and then manually removing
all in-cell commas, so that we can create a transposed_master elsewhere in the
application, because that's a major pain in the backside. So what we're going
here is performing clean-up functions on the resulting compiled_master.xlsx
file that is created straight after the BICC forms compilation. That will then
absolve us from all hassle.
"""
from datetime import date
import pytest

from openpyxl import Workbook, load_workbook
from bcompiler.process import clean_master, clean
from bcompiler.process import Cleanser


@pytest.fixture
def dirty_master():
    inc = 0
    wb = Workbook()
    ws = wb.active
    header_row = ws['A1':'E1']
    data_row = ws['A2':'E2']
    comma_row = ws['A3':'E3']
    newline_row = ws['A4':'E4']
    apostrophe_row = ws['A5':'E5']

    # we need data that isn't set by a loop for row 6
    ws['A6'].value = '20/1/75'
    ws['B6'].value = '21/10/16'
    ws['C6'].value = '2/1/2016'
    ws['D6'].value = '8/7/2220'
    ws['E6'].value = '18/7/20'

    integer_row = ws['A7':'E7']

    for cell in header_row[0]:
        cell.value = "Header {}".format(inc)
        inc += 1
    inc = 0
    for cell in data_row[0]:
        cell.value = "Data {}".format(inc)
        inc += 1
    # now we're going to put some horrible commas in these cells
    for cell in comma_row[0]:
        cell.value = "Garbage data, with commas!"
    for cell in newline_row[0]:
        cell.value = "Garbage data with\nnewlines!"
    for cell in apostrophe_row[0]:
        cell.value = "'Apostrophe! The pernicious apostrophe."
    inc = 2313233
    for cell in integer_row[0]:
        cell.value = str(inc)
        inc += 1

    # we need data that isn't set by a loop for row 8
    # this is for negative integers
    ws['A8'].value = '-232'
    ws['B8'].value = '-233222'
    ws['C8'].value = '-9'
    ws['D8'].value = '-87'
    ws['E8'].value = '-9'

    # we need data that isn't set by a loop for row 9
    # this if for floats
    ws['A9'].value = '2.115'
    ws['B9'].value = '222001.115'
    ws['C9'].value = '0.15'
    ws['D9'].value = '80.99'
    ws['E9'].value = '9.115'

    # we need data that isn't set by a loop for row 10
    # this if for negative floats
    ws['A10'].value = '-2.115'
    ws['B10'].value = '-222001.115'
    ws['C10'].value = '-0.15'
    ws['D10'].value = '-80.99'
    ws['E10'].value = '-9.115'
    return wb


def test_write_wb(dirty_master):
    dirty_master.save('/tmp/dirty_master.xlsx')


def test_phantom_wb(dirty_master):
    """
    Just test that we are creating an openpyxl object for testing. Thanks
    pytest!
    """
    ws = dirty_master.active
    assert ws['A1'].value == 'Header 0'
    assert ws['B1'].value == 'Header 1'
    assert ws['C1'].value == 'Header 2'


def test_presence_of_garbage(dirty_master):
    ws = dirty_master.active
    assert ',' in ws['C3'].value
    assert 'X' not in ws['C3'].value
    assert 'Garbage' in ws['C3'].value
    assert '\n' in ws['A4'].value
    assert 'with\n' in ws['A4'].value


def test_clean_string():
    sample_date = date(1975, 1, 24)
    comma_str = clean('Bobbins, there is nothing here!')
    assert comma_str == 'Bobbins there is nothing here!'
    nl_str = clean('Totally\nrubbish!')
    assert nl_str == 'Totally | rubbish!'
    ap_str = clean("'There is Johnnys cobbles")
    assert ap_str == 'There is Johnnys cobbles'
    d_str = clean('24/01/75')
    assert isinstance(d_str, type(sample_date))
    # it seems that dateuil library defaults to 2065 if format is "DD/MM/65"
    # if "DD/MM/66" however, it will give you 1966
    d_pre_65 = clean('24/1/65')
    assert d_pre_65.year == 2065
    d_post_65 = clean('24/1/66')
    assert d_post_65.year == 1966
    int_str = "12339"
    assert clean(int_str) == 12339
    neg_int_str = "-2333"
    assert clean(neg_int_str) == -2333
    fl_str = "12.23"
    assert clean(fl_str) == 12.23
    neg_fl_str = "-12.23"
    assert clean(neg_fl_str) == -12.23
    removed_bull_dirt = "Chunky\n•Ballcocks"
    assert clean(removed_bull_dirt) == "Chunky | Ballcocks"


@pytest.mark.skip(reason='''For some reason we can't handle ValueError''')
def test_bad_date():
    bad_date_str = '24/0/75'
    with pytest.raises(ValueError):
        clean(bad_date_str)


def test_clean_master(dirty_master):
    # we get this from the fixture: the auto-gen workbook
    dirty_ws = dirty_master.active
    # we need to pass the path of the workbook to kill_commas()
    # this is simulated, because it doesn't have a path as it was
    # generated in the fixture
    path = '/tmp/dirty_master.xlsx'
    # when clean_master() runs, it outputs the workbook it is given
    # with '_cleaned' appended
    c_path = '/tmp/dirty_master_cleaned.xlsx'
    # give it the openpyxl wb object and give it the sheet and the path
    # of the openpyxl object
    clean_master(dirty_master, dirty_ws.title, path)
    # now clean_master() is done, we expect to find a cleaned xlsx file
    cleaned_wb = load_workbook(c_path)
    cleaned_ws = cleaned_wb.active
    sample_date = date(2012, 12, 21)
    assert ',' not in cleaned_ws['C3'].value
    assert cleaned_ws['C3'].value == 'Garbage data with commas!'
    assert '\n' not in cleaned_ws['A4'].value
    assert cleaned_ws['A4'].value == "Garbage data with | newlines!"
    assert cleaned_ws['A5'].value == "Apostrophe! The pernicious apostrophe."
    assert isinstance(cleaned_ws['A6'].value, type(sample_date))
    assert cleaned_ws['B7'].value == 2313234
    assert cleaned_ws['C8'].value == -9
    assert cleaned_ws['D8'].value == -87
    assert cleaned_ws['A9'].value == 2.115
    assert cleaned_ws['A10'].value == -2.115


def test_cleanser_class():
    # comma strings
    commas_str = ("There is tonnes of stuff to think about, we need to clean."
                  " There are multiple commas in here, see? Big commas, big!")
    commas_str2 = ("Millions, upon, millions, of commas! We love ,commas"
                   " even,  if they are malplaced, okay?? , ")

    # apostrophe strings
    apos_str = "'Bobbins ' ' ' ''"
    apos_str2 = "Bobbins ' ' ' ''"

    # mix apos and comma strings
    mix_apos_commas = "'There are mixes, here! Aren't there, yes!"

    # newline strings
    newline_str1 = "There are many ways to write newlines\nand this is one."
    newline_str2 = "Bobbins\nbobbins\nbobbins\nbobbins\nbobbins"

    # date strings
    # dd/mm/yyyy format
    d_d_str = "03/06/2017"

    # with 0:00:00 time format
    d_time_str = "2015-04-01 0:00:00"
    d_time_str2 = "2015-12-31 0:00:00"
    d_time_bad_date = "2015-12-32 0:00:00"

    # integer strings
    i_str = "1234"

    # float strings
    f_str = "12.34"

    # create Cleanser objects for them all
    c = Cleanser(commas_str)
    c2 = Cleanser(commas_str2)
    a = Cleanser(apos_str)
    a2 = Cleanser(apos_str2)
    mix = Cleanser(mix_apos_commas)
    nl = Cleanser(newline_str1)
    nl2 = Cleanser(newline_str2)
    d = Cleanser(d_d_str)
    dt = Cleanser(d_time_str)
    dt2 = Cleanser(d_time_str2)
    d_bad_date = Cleanser(d_time_bad_date)
    i = Cleanser(i_str)
    f = Cleanser(f_str)

    # testing private interface to ensure counting of targets is done
    assert c._checks[c._access_checks('commas')]['count'] == 3
    assert c2._checks[c._access_checks('commas')]['count'] == 7
    assert a._checks[c._access_checks('leading_apostrophe')]['count'] == 1
    assert a2._checks[c._access_checks('leading_apostrophe')]['count'] == 0
    assert mix._checks[c._access_checks('commas')]['count'] == 2
    assert mix._checks[c._access_checks('leading_apostrophe')]['count'] == 1

    # regex checks
    assert nl2.clean() == "Bobbins | bobbins | bobbins | bobbins | bobbins"
    assert nl.clean() == ("There are many ways to write newlines | and this "
                          "is one.")
    assert mix.clean() == "There are mixes here! Aren't there yes!"
    assert c.clean() == ("There is tonnes of stuff to think about we need "
                         "to clean. There are multiple commas in here see? "
                         "Big commas big!")
    assert c2.clean() == ("Millions upon millions of commas! We love "
                          "commas even if they are malplaced okay?? ")
    assert dt.clean() == date(2015, 4, 1)
    assert dt2.clean() == date(2015, 12, 31)
    assert d_bad_date.clean() == d_time_bad_date  # return it and log error
    assert d.clean().month == 6
    assert d.clean().year == 2017
    assert d.clean().day == 3
    assert i.clean() == 1234
    assert f.clean() == 12.34

    # TODO - unable to detect strings in "2015-02-23" format as yet
#    assert d2.clean().month == 6
#    assert d2.clean().year == 2017
#    assert d2.clean().day == 3
