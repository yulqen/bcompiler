import bcompiler.main as main_module
import configparser
import glob
import os

from openpyxl import load_workbook

from ..main import populate_blank_bicc_form as populate
from ..main import get_list_projects
from ..utils import project_data_from_master


config = configparser.ConfigParser()
CONFIG_FILE = 'test_config.ini'
config.read(CONFIG_FILE)


def test_get_list_projects_main_xlsx(master):
    l = get_list_projects(master)
    assert l[0] == 'PROJECT/PROGRAMME NAME 1'


def test_pull_data_from_xlsx_master(master):
    data = project_data_from_master(master)
    assert data['PROJECT/PROGRAMME NAME 1']['SRO Sign-Off'] == 'SRO SIGN-OFF 1'
    assert data['PROJECT/PROGRAMME NAME 1'][
        'Reporting period (GMPP - Snapshot Date)'] == 'REPORTING PERIOD (GMPP - SNAPSHOT DATE) 1'


def test_populate_single_template(master, blank_template):
    SOURCE_DIR = '/tmp/bcompiler-test'
    OUTPUT_DIR = '/tmp/bcompiler-test-output/'
    setattr(main_module, 'OUTPUT_DIR', OUTPUT_DIR)
    setattr(main_module, 'SOURCE_DIR', SOURCE_DIR)
    setattr(main_module, 'BLANK_TEMPLATE_FN', ''.join(['/', blank_template.split('/')[-1]]))
    populate(master, 0)
    wb = load_workbook(os.path.join(OUTPUT_DIR, 'PROJECT_PROGRAMME NAME 1_Q2 Jul - Oct 2017_Return.xlsm'))
    ws = wb[config['TemplateTestData']['summary_sheet']]
    assert ws['B5'].value == 'PROJECT/PROGRAMME NAME 1'
    for f in glob.glob('/'.join([OUTPUT_DIR, '*_Return.xlsm'])):
        os.remove(f)


def test_populate_date_format(master, blank_template):
    SOURCE_DIR = '/tmp/bcompiler-test'
    OUTPUT_DIR = '/tmp/bcompiler-test-output/'
    setattr(main_module, 'OUTPUT_DIR', OUTPUT_DIR)
    setattr(main_module, 'SOURCE_DIR', SOURCE_DIR)
    setattr(main_module, 'BLANK_TEMPLATE_FN', ''.join(['/', blank_template.split('/')[-1]]))
    populate(master, 0)
    wb = load_workbook(os.path.join(OUTPUT_DIR, 'PROJECT_PROGRAMME NAME 1_Q2 Jul - Oct 2017_Return.xlsm'))
    ws = wb[config['TemplateTestData']['fb_sheet']]
    assert ws['E12'].value == '20/06/2017'
    for f in glob.glob('/'.join([OUTPUT_DIR, '*_Return.xlsm'])):
        os.remove(f)
