import csv
import os
import unittest

from shutil import copyfile

from bcompiler.templates import CommissioningTemplate
from bcompiler.main import populate_blank_bicc_form, clean_datamap
from bcompiler.main import parse_csv_to_file
from bcompiler.utils import DATAMAP_MASTER_TO_RETURN
from bcompiler.pipelines.master_returns import create_master_dict_transposed

from openpyxl import load_workbook


class TestForColourFormatting(unittest.TestCase):
    """
    We have problems with the written Excel file having changed cell
    colours. We need to ensure they are correct after writing.
    """
    def setUp(self):
        # we're not going to write the output file into the output folder
        # we're going to write it to output, copy it tmp and then delete it!
        self.tmp_source_path = os.path.abspath('/tmp/')
        # the corporate green colour
        self.known_cell_colour = "40998a"
        self.docs = os.path.join(
            os.path.expanduser('~'), 'Documents')
        bcomp_working_d = 'bcompiler'
        self.path = os.path.join(
            self.docs, bcomp_working_d)
        self.output_path = os.path.join(
            self.path, 'output')
        self.source_path = os.path.join(
            self.path, 'source')
        self.master = os.path.join(
            self.source_path, 'master.csv')
        self.bicc_template = os.path.join(
            self.source_path, 'bicc_template.xlsx')
        # output the file which is at column 2 of the master
        clean_datamap(DATAMAP_MASTER_TO_RETURN)
        parse_csv_to_file(self.master)
        # FIXME this will only work if column 1 of master is this
        # particular project
        populate_blank_bicc_form(self.master, 1)
        copyfile(os.path.join(
            self.output_path, (
                'South Eastern Rail Franchise_Q2_Return.xlsx')), (self.tmp_source_path) + '/test.xlsx')

    @unittest.skip("need to fix a bug before we can run")
    def test_colors(self):
        wb = load_workbook('/tmp/test.xlsx')
        ws = wb['Summary']
        c = ws['A5']
        # FIXME - cannot run this test at the moment because I seemingly
        # cannot load_workbook() a file I have just processed, which is a
        # necessary action as part of this test.
        pass


class TestCommissioningTemplate(unittest.TestCase):
    def setUp(self):
        self.docs = os.path.join(
            os.path.expanduser('~'), 'Documents')
        bcomp_working_d = 'bcompiler'
        self.path = os.path.join(
            self.docs, bcomp_working_d)
        self.source_path = os.path.join(
            self.path, 'source')
        self.output_path = os.path.join(
            self.path, 'output')
        self.datamap_master_to_returns = os.path.join(
            self.source_path, 'datamap-master-to-returns')
        self.datamap_returns_to_master = os.path.join(
            self.source_path, 'datamap-returns-to-master')
        self.master = os.path.join(
            self.source_path, 'master.csv')
        self.transposed_master = os.path.join(
            self.source_path, 'master_transposed.csv')
        self.example_return = os.path.join(
            self.source_path, 'returns/SARH_Q2_Return_Final.xlsx')
        self.bicc_template = os.path.join(
            self.source_path, 'bicc_template.xlsx')
        # test for this, but we want this for multiple tests
        self.ct = CommissioningTemplate()

    def test_presence_of_blank_bicc_template_file(self):
        self.assertTrue(os.path.exists(self.bicc_template))

    def test_for_commissioning_template_object(self):
        commissioning_template_blank = CommissioningTemplate()
        self.assertIsInstance(
            commissioning_template_blank,
            CommissioningTemplate)

    def test_for_commissioning_template_basic_data(self):
        sheets = self.ct.sheets
        self.assertIn('Summary', sheets)
        self.assertIn('Resources', sheets)
        self.assertNotIn('Non Sheet', sheets)
        self.assertIn('Approval & Project milestones', sheets)
        self.assertIn('Assurance planning', sheets)
        self.assertIn('Dropdown List', sheets)

    def test_for_blank_template(self):
        """
        The template itself, prior to commissioning should have no data
        in it, so this test picks a load of random cells and checks that their
        contents is empty. Here we are looking at random cells across the
        workbook.
        """
        summary_sheet = self.ct.openpyxl_obj['Summary']
        resources_sheet = self.ct.openpyxl_obj['Resources']
        self.assertTrue(self.ct.blank)
        self.assertEqual(summary_sheet['B5'].value, None)
        self.assertEqual(summary_sheet['B6'].value, None)
        self.assertEqual(summary_sheet['B8'].value, None)
        self.assertEqual(summary_sheet['C20'].value, None)
        self.assertEqual(summary_sheet['B37'].value, None)
        self.assertEqual(summary_sheet['F41'].value, 'Project Classification')
        self.assertEqual(resources_sheet['E11'].value, None)
        self.assertEqual(resources_sheet['A23'].value, 'Function / Expertise')

    def test_presence_base_master_csv(self):
        self.assertTrue(os.path.exists(self.master))

    def test_presence_transposed_csv(self):
        create_master_dict_transposed(self.master)
        self.assertTrue(os.path.exists(self.transposed_master))

    def test_for_individual_project_data_lines(self):
        test_st = ("Project/Programme Name,SRO Sign-Off,"
                   "FD Sign-Off,")
        len_test_st = len(test_st)
        with open(self.transposed_master, 'r') as f:
            f_line = f.readline(len_test_st)
            self.assertEqual(f_line, test_st)

    def test_for_data_migrated_to_blank_form(self):
        selected_data = {}
        with open(self.transposed_master, 'r') as f:
            projects = list([row for row in csv.DictReader(f)])
            project = projects[0]
            selected_data['Project/Programme Name'] = project[
                'Project/Programme Name']
            selected_data['SRO Sign-Off'] = project['SRO Sign-Off']
            selected_data['Change Delivery Methodology'] = project[
                'Change Delivery Methodology']
        # pass for now because this should be done with a fixture
        # TODO - make this happen
        pass

if __name__ == "__main__":
    unittest.main()
