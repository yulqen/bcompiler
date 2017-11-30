from openpyxl import load_workbook
from ..analysers import rcf_run


def test_rcf(master_with_quarter_year_in_filename, tmpdir):
    wb_master = load_workbook(master_with_quarter_year_in_filename)
    wb_master.save(tmpdir.join(master_with_quarter_year_in_filename.split('/')[-1]))
    rcf_run(tmpdir)
    wb = load_workbook(tmpdir.join('PROJECT_PROGRAMME_NAME_1_Q1_17_18_RCF.xlsx'))
    ws = wb.active
    assert ws['B2'].value == "Reporting period (GMPP - Snapshot Date)"
