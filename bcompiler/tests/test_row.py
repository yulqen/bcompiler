from ..core import Row
from openpyxl import Workbook, load_workbook


def test_for_basic_row_object_given_list(tmpdir):
    values_l = ['Test Value A1', 'Test Value B1', 'Test Value C1']
    wb = Workbook()
    ws = wb.active
    r = Row(1, 1, values_l)
    r.bind(ws)
    wb.save(tmpdir.join('test_row_object.xlsx'))
    loaded_wb = load_workbook(tmpdir.join('test_row_object.xlsx'))
    ws = loaded_wb.active
    assert ws['A1'].value == 'Test Value A1'
    assert ws['B1'].value == 'Test Value B1'
    assert ws['C1'].value == 'Test Value C1'
    assert ws['D1'].value is None


def test_row_object_given_column_reference_as_string(tmpdir):
    values_l = ['Test Value A1', 'Test Value B1', 'Test Value C1']
    wb = Workbook()
    ws = wb.active
    r = Row('A', 1, values_l)
    r.bind(ws)
    wb.save(tmpdir.join('test_row_object.xlsx'))
    loaded_wb = load_workbook(tmpdir.join('test_row_object.xlsx'))
    ws = loaded_wb.active
    assert ws['A1'].value == 'Test Value A1'
    assert ws['B1'].value == 'Test Value B1'
    assert ws['C1'].value == 'Test Value C1'
    assert ws['D1'].value is None


def test_row_object_given_column_reference_as_double_string(tmpdir):
    values_l = ['Test Value AA1', 'Test Value AB1', 'Test Value AC1']
    wb = Workbook()
    ws = wb.active
    r = Row('AA', 1, values_l)
    r.bind(ws)
    wb.save(tmpdir.join('test_row_object.xlsx'))
    loaded_wb = load_workbook(tmpdir.join('test_row_object.xlsx'))
    ws = loaded_wb.active
    assert ws['AA1'].value == 'Test Value AA1'
    assert ws['AB1'].value == 'Test Value AB1'
    assert ws['AC1'].value == 'Test Value AC1'
