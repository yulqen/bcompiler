import datetime

from openpyxl import load_workbook

from ..analysers.swimlane import run as swimlane_run


def test_basic_swimlane_data(tmpdir, master):
    tmpdir = [tmpdir]  # hacking the fact that output_path in implementation is list
    swimlane_run(output_path=tmpdir, user_provided_master_path=master)
    tmpdir = tmpdir[0]  # hacking the fact that output_path in implementation is list
    output = load_workbook(tmpdir.join('swimlane_milestones.xlsx'))
    ws = output.active
    assert ws['A1'].value == "PROJECT/PROGRAMME NAME 1"
    assert ws['A2'].value == "APPROVAL MM1 1" # config.ini: block_start row
    assert ws['A3'].value == "APPROVAL MM2 1"  # config.ini: block_start + block_skip
    assert ws['A4'].value == "APPROVAL MM3 1"  # config.ini: last one + block_skip
    assert ws['A5'].value == "APPROVAL MM4 1"  # config.ini: last one + block_skip
    assert ws['A6'].value == "APPROVAL MM5 1"  # config.ini: last one + block_skip
    assert ws['A7'].value == "APPROVAL MM6 1"  # config.ini: last one + block_skip
    assert ws['A8'].value == "APPROVAL MM7 1"  # config.ini: last one + block_skip
    assert ws['A9'].value == "APPROVAL MM8 1"  # config.ini: last one + block_skip
    assert ws['A10'].value == "APPROVAL MM9 1"  # config.ini: last one + block_skip

    assert ws['B2'].value == datetime.datetime(2015, 1, 1)
    assert ws['B3'].value == datetime.datetime(2019, 1, 1)
