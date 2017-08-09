import csv
import fnmatch
import logging
import os
from datetime import date, datetime
from math import isclose

import configparser

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import quote_sheetname

from bcompiler.datamap import DatamapGMPP

logger = logging.getLogger('bcompiler.utils')

rdel_cdel_merge = ''

CURRENT_QUARTER = "Q2 Jul - Sep 2017"

DOCS = os.path.join(os.path.expanduser('~'), 'Documents')
BCOMPILER_WORKING_D = 'bcompiler'
ROOT_PATH = os.path.join(DOCS, BCOMPILER_WORKING_D)

CONFIG_FILE = os.path.join(DOCS, BCOMPILER_WORKING_D, 'config.ini')


config = configparser.ConfigParser()
config.read(CONFIG_FILE)

SHEETS = [i for i in dict((config.items('TemplateSheets'))).values()]

def quick_typechecker(*args):
    """
    Very simple function to filter allowed types (int, float). Any other type
    returns False. All arguments must be of same type.
    """
    for arg in args:
        if isinstance(arg, (int, float, date)):
            pass
        else:
            return False
    return True


def simple_round(fl, prec):
    """Rounds a fl to prec precision."""
    return round(fl, prec)


def bc_is_close(x, y):
    """Returns true if acceptably close."""
    if isinstance(x, (date, datetime)) or isinstance(y, (date, datetime)):
        return False
    else:
        return isclose(x, y, rel_tol=0.001)


def cell_bg_colour(rgb=[]):
    """
    Give it a list of integers between 0 and 255 - three of them.
    """
    c_value = "{0:02X}{1:02X}{2:02X}".format(*rgb)
    return PatternFill(patternType='solid', fgColor=c_value, bgColor=c_value)


def get_relevant_names(project_name, project_data):

    try:
        sro_first_name = project_data[project_name]['SRO Full Name'].split(
            " ")[0]
    except IndexError:
        logger.warning("SRO Full Name ({0}) is not suitable for splitting".
                       format(project_data[project_name]['SRO Full Name']))

    try:
        sro_last_name = project_data[project_name]['SRO Full Name'].split(" ")[
            1]
    except IndexError:
        logger.warning("SRO Full Name ({0}) is not suitable for splitting".
                       format(project_data[project_name]['SRO Full Name']))

    try:
        pd_first_name = project_data[project_name]['PD Full Name'].split(" ")[
            0]
    except IndexError:
        logger.warning("PD Full Name ({0}) is not suitable for splitting".
                       format(project_data[project_name]['PD Full Name']))

    try:
        pd_last_name = project_data[project_name]['PD Full Name'].split(" ")[1]
    except IndexError:
        logger.warning("PD Full Name ({0}) is not suitable for splitting".
                       format(project_data[project_name]['PD Full Name']))

    try:
        sro_d = dict(first_name=sro_first_name, last_name=sro_last_name)
    except UnboundLocalError:
        sro_d = None
    try:
        pd_d = dict(first_name=pd_first_name, last_name=pd_last_name)
    except UnboundLocalError:
        pd_d = None

    return (sro_d, pd_d)


def populate_blank_gmpp_form(openpyxl_template, project):
    blank = openpyxl_template
    dm = DatamapGMPP(
        '/home/lemon/Documents/bcompiler/source/datamap-master-to-gmpp')
    logger.info("Grabbing GMPP datamap {}".format(dm.source_file))
    target_ws = blank['GMPP Return']
    project_data = project_data_line()

    relevant_names = get_relevant_names(project, project_data)
    if relevant_names[0] and relevant_names[1]:
        relevant_names = get_relevant_names(project, project_data)
    else:
        relevant_names = [({
            'first_name': '',
            'last_name': ''
        }), ({
            'first_name': '',
            'last_name': ''
        })]

    for line in dm.data:

        if 'Project/Programme Name' in line.cellname:
            d_to_migrate = project
            target_ws[line.cellref].value = d_to_migrate

        elif line.cellref is not None:
            if line.cellname == 'SRO First Name':
                d_to_migrate = relevant_names[0]['first_name']
                target_ws[line.cellref].value = d_to_migrate
            if line.cellname == 'SRO Last Name':
                d_to_migrate = relevant_names[0]['last_name']
                target_ws[line.cellref].value = d_to_migrate
            if line.cellname == 'PD First Name':
                d_to_migrate = relevant_names[1]['first_name']
                target_ws[line.cellref].value = d_to_migrate
            if line.cellname == 'PD Last Name':
                d_to_migrate = relevant_names[1]['last_name']
                target_ws[line.cellref].value = d_to_migrate

            try:
                # pull the data if we can
                d_to_migrate = project_data[project][line.cellname]
            except KeyError:
                logger.warning(("Unable to find {} in master intended for {}"
                                " in template").format(line.cellname,
                                                       line.cellref))
            else:
                target_ws[line.cellref].value = d_to_migrate
                logger.debug("Migrating {} from {} to blank template".format(
                    d_to_migrate, line.cellref))
    # inject additonal data
    additional_data = dm.add_additional_data()
    for line in additional_data:
        target_ws[line.cellref].value = line.added_data_field
    fn = OUTPUT_DIR + project + ' Q1_GMPP.xlsx'
    logger.info("Writing {}".format(fn))
    blank.save(fn)


def project_data_line():
    p_dict = {}
    with open(SOURCE_DIR + 'master_transposed.csv', 'r') as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = row.pop('Project/Programme Name')
            if key in p_dict:
                pass
            p_dict[key] = row
            logger.debug(
                "Adding {} to project_data_line dictionary".format(key))
    return p_dict


def gmpp_project_data():
    data = project_data_line()
    gmpp_project_data_list = []
    for project in data:
        if data[project]['GMPP (GMPP - formally joined GMPP)']:
            gmpp_project_data_list.append(data[project])
    return gmpp_project_data_list


def gmpp_project_names():
    data = project_data_line()
    return [
        project for project in data
        if data[project]['GMPP (GMPP - formally joined GMPP)'] != "No" and
        data[project]['GMPP (GMPP - formally joined GMPP)'] != "NA" and data[
            project]['GMPP (GMPP - formally joined GMPP)'] != ""
    ]


def open_openpyxl_template(template_file):
    """
    Opens an xlsx file (the template) and returns the openpyxl object.
    """
    wb = load_workbook(template_file, keep_vba=True)
    logger.info("Opening {} as an openpyxl object".format(template_file))
    return wb


def working_directory(dir_type=None):
    """
    Returns the working direct for source files
    :return: path to the working directory intended for the source files
    """
    docs = os.path.join(os.path.expanduser('~'), 'Documents')
    try:
        bcomp_working_d = 'bcompiler'
    except FileNotFoundError:
        print("You need to run with --create-wd to",
              "create the working directory")
    root_path = os.path.join(docs, bcomp_working_d)
    if dir_type == 'source':
        return root_path + "/source/"
    elif dir_type == 'output':
        return root_path + "/output/"
    elif dir_type == 'returns':
        return root_path + "/source/returns/"
    else:
        return


SOURCE_DIR = working_directory('source')
OUTPUT_DIR = working_directory('output')
RETURNS_DIR = working_directory('returns')
DATAMAP_RETURN_TO_MASTER = SOURCE_DIR + 'datamap.csv'
DATAMAP_MASTER_TO_RETURN = SOURCE_DIR + 'datamap.csv'
DATAMAP_MASTER_TO_GMPP = SOURCE_DIR + 'archive/datamap-master-to-gmpp'
CLEANED_DATAMAP = SOURCE_DIR + 'cleaned_datamap.csv'
MASTER = SOURCE_DIR + 'master.csv'
TEMPLATE = SOURCE_DIR + 'bicc_template.xlsx'
GMPP_TEMPLATE = SOURCE_DIR + 'archive/gmpp_template.xlsx'


def index_returns_directory():
    """
    Prior to compiling a master, it is useful to get the order of projects
    by their file name, as the compile.run() function traverses the directory
    top to bottom to build the master. We can then use this to compare with the
    order or projects (columns) in the old master document we are comparing
    the current compile. This is pretty hackish but needs must...
    """
    target_files = []
    for f in os.listdir(RETURNS_DIR):
        target_files.append(f)

    pnames_in_returns_dir = []
    for f in target_files:
        if fnmatch.fnmatch(f, '*.xlsx'):
            wb = load_workbook(os.path.join(RETURNS_DIR, f))
            ws = wb['Summary']
            pnames_in_returns_dir.append(ws['B5'].value)
    return pnames_in_returns_dir


def transpose_master_xlsx(source_file):
    """
    Transposes the xlsx master to a new master_transposed.csv file.
    """
    output = open(SOURCE_DIR + 'master_transposed.csv', 'w+')
    wb = load_workbook(source_file)
    ws = wb.active
    pass

def splat_rows(row):
      yield [(c.value, c.row, c.column) for c in row]


def parse_csv_to_file(source_file):
    """
    Transposes the master to a new master_transposed.csv file.
    :param source_file:
    :return:
    """
    output = open(SOURCE_DIR + 'master_transposed.csv', 'w+')
    with open(source_file, 'r') as source_f:
        lis = [x.split(',') for x in source_f]
        for i in lis:
            # we need to do this to remove trailing "\n" from the end of
            # each original master.csv line
            i[-1] = i[-1].rstrip()

    for x in zip(*lis):
        for y in x:
            output.write(y + ',')
        output.write('\n')
    output.close()


def create_master_dict_transposed(source_master_csv):
    """
    The side-effect of the following function is to ensure there is a
    'master_transposed.csv' file present in SOURCE_DIR
    returns a list of dicts, which makes up all the data from the master
    """
    parse_csv_to_file(source_master_csv)
    with open(SOURCE_DIR + 'master_transposed.csv', 'r') as f:
        r = csv.DictReader(f)
        ls = [row for row in r]
    return ls


sheet_name = "Dropdown"

VALIDATION_REFERENCES = {
    'Quarter':
    "{0}!$A$2:$A$9".format(quote_sheetname(sheet_name)),
    'Joining Qtr':
    "{0}!$B$2:$B$25".format(quote_sheetname(sheet_name)),
    'Classification':
    "{0}!$C$2:$C$4".format(quote_sheetname(sheet_name)),
    'Entity format':
    "{0}!$D$2:$D$4".format(quote_sheetname(sheet_name)),
    'Methodology':
    "{0}!$E$2:$E$10".format(quote_sheetname(sheet_name)),
    'Category':
    "{0}!$F$2:$H$11".format(quote_sheetname(sheet_name)),
    'Scope Changed':
    "{0}!$G$2:$I$4".format(quote_sheetname(sheet_name)),
    'Monetised / Non Monetised Benefits':
    "{0}!$H$2:$H$4".format(quote_sheetname(sheet_name)),
    'RAG':
    "{0}!$I$2:$I$6".format(quote_sheetname(sheet_name)),
    'RAG 2':
    "{0}!$J$2:$J$4".format(quote_sheetname(sheet_name)),
    'RPA level':
    "{0}!$K$2:$K$4".format(quote_sheetname(sheet_name)),
    'Capability RAG':
    "{0}!$L$2:$L$5".format(quote_sheetname(sheet_name)),
    'MPLA / PLP':
    "{0}!$M$2:$M$30".format(quote_sheetname(sheet_name)),
    'PL Changes':
    "{0}!$N$2:$N$31".format(quote_sheetname(sheet_name)),
    'Stage':
    "{0}!$O$2:$O$10".format(quote_sheetname(sheet_name)),
    'Business Cases':
    "{0}!$P$2:$P$11".format(quote_sheetname(sheet_name)),
    'Milestone Types':
    "{0}!$Q$2:$Q$4".format(quote_sheetname(sheet_name)),
    'Finance figures format':
    "{0}!$R$2:$R$3".format(quote_sheetname(sheet_name)),
    'Index Years':
    "{0}!$S$2:$S$27".format(quote_sheetname(sheet_name)),
    'Discount Rate':
    "{0}!$T$2:$T$32".format(quote_sheetname(sheet_name)),
    'Finance type':
    "{0}!$U$2:$U$6".format(quote_sheetname(sheet_name)),
    'Yes/No':
    "{0}!$V$2:$V$3".format(quote_sheetname(sheet_name)),
    'Years (Spend)':
    "{0}!$W$2:$W$90".format(quote_sheetname(sheet_name)),
    'Years (Benefits)':
    "{0}!$X$2:$X$90".format(quote_sheetname(sheet_name)),
    'Snapshot Dates':
    "{0}!$Y$2:$Y$9".format(quote_sheetname(sheet_name)),
    'Percentage of time spent on SRO role':
    "{0}!$Z$2:$Z$21".format(quote_sheetname(sheet_name)),
    'AR Category':
    "{0}!$AA$2:$AA$5".format(quote_sheetname(sheet_name)),
    'Project Lifecycle':
    "{0}!$AB$2:$AB$6".format(quote_sheetname(sheet_name)),
    'Programme Lifecycle':
    "{0}!$AC$2:$AC$7".format(quote_sheetname(sheet_name)),
    'Other':
    "{0}!$AD$2:$AD$19".format(quote_sheetname(sheet_name)),
    'Start / Year end - FY':
    "{0}!$AE$3:$AE$22".format(quote_sheetname(sheet_name)),
    'Count':
    "{0}!$AF$2:$AF$22".format(quote_sheetname(sheet_name)),
    'VFM':
    "{0}!$AG$2:$AG$11".format(quote_sheetname(sheet_name)),
    'DfT Group':
    "{0}!$AH$2:$AH$7".format(quote_sheetname(sheet_name)),
    'DfT Division':
    "{0}!$AI$2:$AI$15".format(quote_sheetname(sheet_name)),
    'Agency':
    "{0}!$AJ$2:$AJ$9".format(quote_sheetname(sheet_name)),
    'High Speed Rail':
    "{0}!$AK$2:$AK$4".format(quote_sheetname(sheet_name)),
    'Rail Group':
    "{0}!$AL$2:$AL$4".format(quote_sheetname(sheet_name)),
    'Roads, Devolution & Motoring':
    "{0}!$AM$2:$AM$5".format(quote_sheetname(sheet_name)),
    'International, Security and Environment':
    "{0}!$AN$2:$AN$4".format(quote_sheetname(sheet_name)),
    'Resource and Strategy':
    "{0}!$AO$2:$AO$2".format(quote_sheetname(sheet_name)),
    'Non-Group':
    "{0}!$AP$2:$AP$2".format(quote_sheetname(sheet_name)),
    'GMPP Annual Report Category':
    "{0}!$AQ$2:$AQ$2".format(quote_sheetname(sheet_name)),
    'SDP':
    "{0}!$AR2:$AR$5".format(quote_sheetname(sheet_name)),
}



def row_accessor(row: tuple):
    """
    Utility generator yielding tuple of form (str, str); e.g
    ('A10', 'Project/Programme Name').
    :param row:
    :return:
    """
    for item in row:
        yield (''.join([item.column, str(item.row)]), item.value)


def gen_sheet_data(workbook: str) -> dict:
    """
    Returns a dict containing data from a given xlsx file, by sheet
    within that workbook.
    :param path to xlsx file:
    :return: dict of data by sheet in workbook
    """
    wb = load_workbook(workbook)
    sheets = wb._sheets
    data = {}
    for s in sheets:
        rows = s.rows
        title = s.title
        data[title] = [list(row_accessor(x)) for x in rows]
    return data


def parse_data_row(row: list) -> tuple:
    """
    Utility generator which processes two-item tuples in a list.
    :param row:
    :return: tuple of form (str, str); e.g. ('A10', 'Project/Programme Name')
    """
    for item in row:
        yield item[0], item[1]


def get_sheets_in_workbook(real_template: str) -> list:
    """
    Utility function to return a list of sheet names from an xlsx file.
    :param real_template:
    :return: list of sheet names
    """
    wb = load_workbook(real_template)
    sheets = wb._sheets
    return sheets


def generate_test_template_from_real(real_template: str, save_path: str) -> None:
    """
    Given the bicc_template.xlsx file, this function strips it of
    everything but cell data.
    :param real_template: str path of location of bicc_template.xlsx
    :param save_path: str path of output directory; file will be named 'gen_bicc_template.xlsx',
    of the form "~/Documents"
    :return:
    """
    data = gen_sheet_data(real_template)
    sheets = get_sheets_in_workbook(real_template)
    blank = Workbook()
    sheet_order = 0
    for sheet in sheets:
        summary_sheet = blank.create_sheet(sheet.title, sheet_order)
        for row in data[sheet.title]:
            r = parse_data_row(row)
            for cell in r:
                summary_sheet[cell[0]] = cell[1]
        sheet_order += 1
    if save_path.endswith('/'):
        save_path = save_path[:-1]
    blank.save(''.join([save_path, '/gen_bicc_template.xlsx']))


