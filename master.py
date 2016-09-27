from bcmaster import BCMasterCSV
from datamap import DataMap
from openpyxl import load_workbook, Workbook


m = BCMasterCSV('source_files/master.csv', as_dataframe=True)
#m = m.flip()
m_dict = m.as_dataframe.to_dict()
list_of_projects = m.as_dataframe.T.index
dict_keys = m_dict.keys()


dm = DataMap('source_files/datamap')
dm.parse()

# let's open a blank template

project1 = list_of_projects[10]
wb = load_workbook('source_files/bicc_template.xlsx')
ws_summary = wb.get_sheet_by_name("Summary")
ws_fb = wb.get_sheet_by_name("Finance & Benefits")

project1_data = m_dict[project1]
project1_data['Project/Programme Name'] = project1

#print(ws['A5'].value)
#ws['B5'] = project1_data['Project/Programme Name']
#print("Successfully wrote data")
#wb.save('source_files/test1.xlsx')

for item in dm.output_excel_map_list:
    if item['sheet'] == 'Summary':
        try:
            ws_summary[item['cell_coordinates']].value = project1_data[item['cell_description']]
        except KeyError:
            print("Cannot find {} in master.csv".format(item['cell_description']))
            pass
    elif item['sheet'] == 'Finance & Benefits':
        try:
            ws_fb[item['cell_coordinates']].value = project1_data[item['cell_description']]
        except KeyError:
            print("Cannot find {} in master.csv".format(item['cell_description']))
            pass

wb.save('source_files/test1.xlsx')

