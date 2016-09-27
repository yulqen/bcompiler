from openpyxl import load_workbook

wb = load_workbook('/home/lemon/code/python/bicc_excel/source_files/bicc_template.xlsx')
ws = wb.get_sheet_by_name('Summary')
print(ws['A5'].value)
ws['B5'] = 'A Funny Project'
ws['B6'] = 'GIBBERISH'
ws['B8'] = 'A Group of People Who Care!'
wb.save('/home/lemon/code/python/bicc_excel/source_files/bicc_template_completed.xlsx')
print('Written')
