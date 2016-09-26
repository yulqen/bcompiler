import fnmatch
import os
import re
"""
PROBABLY NEVER GOING TO UNDERSTAND THIS COME SEPTEMBER
"""
from openpyxl import load_workbook, Workbook

cell_regex = re.compile('[A-Z]+[0-9]+')

SOURCE_RETURN_FILE = "source_files/A14 Cambs to Huntington.xlsx"
DATA_MAP_FILE = 'source_files/data_map'
OUTPUT_FILE = 'DfT_GMPP_Return.xlsx'

def get_sheet_names(source_file):
    wb = load_workbook(source_file, read_only=True)
    return wb.get_sheet_names()


def get_sheet_data(source_file):
    wb = load_workbook(source_file, read_only=True)
    ws = wb['Finance & Benefits']

    for row in ws.rows:
        for cell in row:
            if cell.value != None:
                print(cell.value)


def get_current_quarter(source_file):
    wb = load_workbook(source_file, read_only=True)
    ws = wb['Summary']
    q = ws['K6'].value
    print(q)


def get_project_name(source_file):
    wb = load_workbook(source_file, read_only=True)
    ws = wb['Summary']
    cn = ws['C10'].value
    print(cn)


def parse_data_file():
    with open(DATA_MAP_FILE, 'r') as f:
        data = f.readlines()

        for line in data:
            words = line.split(',')
            print(words)


def parse_source_cells(source_file):
    """
    :param source_file: an Excel file containing project return data
    :return: a list of dict items mapping each key:value pair for the output column in GMPP's template
    """

    # first, we load the source file
    wb = load_workbook(source_file, read_only=True, data_only=True)

    # we're going to output data from this function as a list of dict items
    output_excel_map_list = []

    # load the DATA_MAP_FILE, containing mappings to cells in the form based on key values
    # from GMPP's master template
    with open(DATA_MAP_FILE, 'r', encoding='UTF-8') as f:
        data = f.readlines()

        for line in data:
            # split on , allowing us to access useful data from data map file
            data_map_line = line.split(',')
            # if the second word in each MAP line is a named sheet from the template file, we're interested
            if data_map_line[1] in ['Summary', 'Finance & Benefits', 'Resources', 'Milestones and Assurance', 'Dropdown lists', 'Resources backup']:
                # the end item in the list is a newline - get rid of that
                del data_map_line[-1]
                # the worksheet in the source Excel file needs to be accessible
                try:
                    ws = wb[data_map_line[1]]
                except KeyError as e:
                    print("{} has no {} sheet! - {}".format(source_file, data_map_line[1], e))
                # we only want to act query the source Excel file if there is a valid cell reference there
                # so we use a regex to do that
                if cell_regex.search(data_map_line[-1]):
                    try:
                        destination_kv = dict(gmpp_key=data_map_line[0], gmpp_key_value=ws[data_map_line[-1]].value)
                    except IndexError:
                        destination_kv = dict(gmpp_key=data_map_line[0], gmpp_key_value="OUT OF BOUNDS!")
                    output_excel_map_list.append(destination_kv)
                    # else:
                    #     destination_kv = dict(gmpp_key=data_map_line[0], gmpp_key_value=[data_map_line[-1][0]])
                    #     output_excel_map_list.append(destination_kv)
            # if the DATA_MAP doesn't suggest the data is sourced in the template Excel, we just want to
            # take the default data we have entered there (e.g. 'michelle dawson' as default)
            # OR we return an empty string if there is no data
            else:
                # the end item in the list is a newline - get rid of that
                del data_map_line[-1]
                if len(data_map_line) > 1:
                    destination_kv = dict(gmpp_key=data_map_line[0], gmpp_key_value=data_map_line[-1])
                # if the list has only one item, that means there is no data entered, so we want the value to
                # be an empty string for now
                else:
                    destination_kv = dict(gmpp_key=data_map_line[0], gmpp_key_value="")
                output_excel_map_list.append(destination_kv)

    return output_excel_map_list


def write_excel(source_file, count, workbook):
    # count is used to count number of times function is run so that multiple returns can be added
    # and not overwrite the GMPP key column
    # let's create an Excel file in memory
    # it will have one worksheet - let's get it
    ws = workbook.active
    # give it a title
    ws.title = "GMPP Return - DfT"

    out_map = parse_source_cells(source_file)

    if count == 1:
        i = 1
        for d in out_map:
            c = ws.cell(row=i, column=1)
            c.value = d['gmpp_key']
            i +=1
        i = 1
        for d in out_map:
            c = ws.cell(row=i, column=2)
            c.value = d['gmpp_key_value']
            i +=1
    else:
        i = 1
        for d in out_map:
            c = ws.cell(row=i, column=count+1)
            c.value = d['gmpp_key_value']
            i +=1

def find_variance(source_file):
    true_map_summary = [
        ('BICC PORTFOLIO OFFICE - GMPP REPORTING RETURN', 'A5'),
        ('PD Tenure Start date', 'H17'),
        ('List Strategic Outcomes (monetised and non-monetised benefits)','A44')
    ]

    wb = load_workbook(('source_files/' + source_file), read_only=True)
    # do check summary page
    ws = wb['Summary']
    for t in true_map_summary:
        target_cell = t[1]
        expected_value = t[0]
        actual_value = ws[target_cell].value
        if expected_value == actual_value:
            print('{}: {} cell {} OK'.format(source_file, ws, target_cell))
        else:
            print('{}: {} cell {} should be {} but is reporting {}'.format(
                source_file,
                ws,
                target_cell,
                expected_value,
                actual_value
            ))




if __name__ == '__main__':
    workbook = Workbook()
    dir = os.path.dirname(os.path.realpath(__file__))
    count = 1
    for file in os.listdir(os.path.join(dir, 'source_files')):
        if fnmatch.fnmatch(file, '*.xlsx'):
            print("Processing {}".format(file))
            write_excel(('source_files/'+file), count=count, workbook=workbook)
            count += 1
    workbook.save(OUTPUT_FILE)

    # find variance
    for file in os.listdir(os.path.join(dir, 'source_files')):
        if fnmatch.fnmatch(file, '*.xlsx'):
            find_variance(file)
