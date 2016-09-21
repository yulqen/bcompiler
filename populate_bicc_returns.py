# -*- coding: utf-8 -*-
"""
    bicc_excel.populate_bicc_returns
    ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    Populate blank BICC Return forms from a master xlsx template.

    :copyright: (c) 2016 by Matthew Lemon.
    
    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.
    
    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.
    
    You should have received a copy of the GNU General Public License
    along with this program.  If not, see <http://www.gnu.org/licenses/>.
    
"""

from openpyxl import load_workbook, Workbook
from time import sleep
import re
import csv
import pandas as pd


cell_regex = re.compile('[A-Z]+[0-9]+')

SOURCE_MASTER = 'q2_source_files/master_test.xlsx'
SOURCE_CSV = 'q2_source_files/master_test.csv'

def get_sheet_names(source_file):
    wb = load_workbook(source_file, read_only=True)
    return wb.get_sheet_names()


def get_sheet_data(source_file):
    wb = load_workbook(source_file, read_only=True)
    ws = wb['GMPP Return - DfT']

    for row in ws.rows:
        for cell in row:
            if cell.value != None:
                print(cell.value)


def get_project_column_data(source_file):
    """This returns a dictionary contain key values for a single project column
    in the master spreadsheet.
    """
    wb = load_workbook(source_file, read_only=True)
    ws = wb['GMPP Return - DfT']
    for row in ws.iter_rows(min_row=1, max_col=2, max_row=847):
        for cell in row:
            print(cell.value)
            sleep(0.5)


# I know - let's convert the damn master into a csv file first, because that's
# what we need!

# First thing we need to do is transpose the rows and columns so we get a
# proper CSV file
#df = pd.read_csv(SOURCE_CSV, index_col=0)
# this does the transposing and writes out to another file
#df.T.to_csv('q2_source_files/another_bicc.csv')
# now we cant to get data for only one project
df = pd.read_csv('q2_source_files/another_bicc.csv', index_col=0)
print(df.loc['A14 Cambridge to Huntingdon Improvement Scheme'].to_dict())
